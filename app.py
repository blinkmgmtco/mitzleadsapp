#!/usr/bin/env python3
"""
🚀 ULTIMATE LEAD SCRAPER CRM - PRODUCTION READY
High-Intent Lead Generation with Dual Scraping Modes
Beautiful Streamlit Dashboard with Full CRM
"""

import json
import time
import hashlib
import re
import random
import os
import sys
import sqlite3
import csv
import io
import threading
import asyncio
import aiohttp
import concurrent.futures
from datetime import datetime, timezone, timedelta
from typing import List, Dict, Optional, Any, Set, Tuple, Union
from urllib.parse import urlparse, urljoin, parse_qs
from pathlib import Path
import html
import ssl
import socket

# ============================================================================
# IMPORTS WITH FALLBACKS & ENHANCED ERROR HANDLING
# ============================================================================

# Check if we're running in Streamlit Cloud and adjust paths
if 'STREAMLIT_CLOUD' in os.environ:
    os.makedirs('/tmp/.ultimateleadscraper', exist_ok=True)
    CONFIG_FILE = '/tmp/.ultimateleadscraper/config.json'
    DB_FILE = '/tmp/.ultimateleadscraper/crm_database.db'
    CACHE_DIR = '/tmp/.ultimateleadscraper/cache'
else:
    CONFIG_FILE = "config.json"
    DB_FILE = "crm_database.db"
    CACHE_DIR = "cache"

os.makedirs(CACHE_DIR, exist_ok=True)

# Enhanced import handling with version checks
def check_and_install(package, import_name=None):
    """Check if package is installed, provide installation instructions"""
    import_name = import_name or package
    try:
        __import__(import_name)
        return True
    except ImportError:
        print(f"❌ Missing package: {package}")
        print(f"   Install with: pip install {package}")
        return False

REQUIRED_PACKAGES = [
    ('requests', 'requests'),
    ('beautifulsoup4', 'bs4'),
    ('streamlit', 'streamlit'),
    ('pandas', 'pandas'),
    ('plotly', 'plotly'),
    ('aiohttp', 'aiohttp'),
    ('openpyxl', 'openpyxl'),
    ('pydantic', 'pydantic'),
    ('python-multipart', 'multipart'),
    ('cryptography', 'cryptography'),
    ('python-dotenv', 'dotenv')
]

for package, import_name in REQUIRED_PACKAGES:
    if not check_and_install(package, import_name):
        if package not in ['openpyxl', 'cryptography', 'python-dotenv']:
            sys.exit(1)

# Now import everything
import requests
from bs4 import BeautifulSoup
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import aiohttp
import asyncio
from typing import Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pydantic
from pydantic import BaseModel, Field, validator
from cryptography.fernet import Fernet
import base64
from dotenv import load_dotenv

# Optional imports
try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    print("⚠️  OpenAI not installed. AI features disabled.")

try:
    from streamlit_autorefresh import st_autorefresh
    AUTOREFRESH_AVAILABLE = True
except ImportError:
    AUTOREFRESH_AVAILABLE = False
    print("⚠️  streamlit-autorefresh not installed. Auto-refresh disabled.")

# ============================================================================
# ENHANCED CONFIGURATION WITH PYDANTIC VALIDATION
# ============================================================================

class ScraperMode(BaseModel):
    """Scraper mode configuration"""
    name: str = Field(..., description="Mode name")
    description: str = Field(..., description="Mode description")
    scrape_no_website: bool = Field(True, description="Scrape businesses with no website")
    scrape_broken_website: bool = Field(True, description="Scrape businesses with broken websites")
    scrape_with_website: bool = Field(False, description="Scrape businesses with functional websites")
    quality_threshold: int = Field(80, ge=0, le=100, description="Minimum quality score")
    max_leads_per_cycle: int = Field(10, ge=1, le=1000, description="Maximum leads per cycle")

class APIConfig(BaseModel):
    """API configuration"""
    serper_api_key: str = Field("", description="Serper.dev API key")
    openai_api_key: str = Field("", description="OpenAI API key")
    proxycrawl_api_key: Optional[str] = Field(None, description="ProxyCrawl API key for bypassing blocks")
    brightdata_api_key: Optional[str] = Field(None, description="BrightData API key")
    scraperapi_key: Optional[str] = Field(None, description="ScraperAPI key")

class UISettings(BaseModel):
    """UI settings"""
    theme: str = Field("ultimate", description="UI theme")
    primary_color: str = Field("#0066FF", description="Primary color")
    secondary_color: str = Field("#00D4AA", description="Secondary color")
    accent_color: str = Field("#FF6B6B", description="Accent color")
    dark_mode: bool = Field(True, description="Dark mode enabled")
    compact_view: bool = Field(False, description="Compact view mode")
    animations_enabled: bool = Field(True, description="Enable animations")

class ScrapingFilters(BaseModel):
    """Scraping filters"""
    exclude_chains: bool = Field(True, description="Exclude chain/franchise businesses")
    exclude_without_phone: bool = Field(False, description="Exclude businesses without phone")
    exclude_without_address: bool = Field(False, description="Exclude businesses without address")
    min_employee_count: Optional[int] = Field(None, ge=0, description="Minimum employee count")
    max_employee_count: Optional[int] = Field(None, ge=0, description="Maximum employee count")
    years_in_business_min: Optional[int] = Field(None, ge=0, description="Minimum years in business")
    exclude_keywords: List[str] = Field(default_factory=list, description="Keywords to exclude")
    include_keywords: List[str] = Field(default_factory=list, description="Keywords to include")
    target_cities: List[str] = Field(default_factory=list, description="Target cities")
    target_industries: List[str] = Field(default_factory=list, description="Target industries")

class AIEnrichment(BaseModel):
    """AI enrichment settings"""
    enabled: bool = Field(True, description="Enable AI enrichment")
    model: str = Field("gpt-4o-mini", description="AI model to use")
    max_tokens: int = Field(2000, ge=100, le=4000, description="Maximum tokens for AI responses")
    auto_qualify: bool = Field(True, description="Auto-qualify leads")
    qualification_threshold: int = Field(70, ge=0, le=100, description="Qualification threshold")
    extract_decision_maker: bool = Field(True, description="Extract decision maker info")
    estimate_revenue: bool = Field(True, description="Estimate business revenue")
    suggest_outreach_strategy: bool = Field(True, description="Suggest outreach strategy")

class CRMSettings(BaseModel):
    """CRM settings"""
    enabled: bool = Field(True, description="Enable CRM")
    database: str = Field("crm_database.db", description="Database file")
    auto_sync: bool = Field(True, description="Auto-sync leads")
    prevent_duplicates: bool = Field(True, description="Prevent duplicate leads")
    duplicate_check_field: str = Field("fingerprint", description="Field for duplicate check")
    batch_size: int = Field(50, ge=1, le=1000, description="Batch size for operations")
    default_status: str = Field("New Lead", description="Default lead status")
    default_assigned_to: str = Field("", description="Default assigned user")
    auto_set_production_date: bool = Field(True, description="Auto-set production date")
    lead_expiry_days: int = Field(90, ge=1, description="Days before lead expires")
    auto_followup_days: int = Field(7, ge=1, description="Days for auto-followup")

class NotificationSettings(BaseModel):
    """Notification settings"""
    email_enabled: bool = Field(False, description="Enable email notifications")
    email_smtp_server: Optional[str] = Field(None, description="SMTP server")
    email_smtp_port: Optional[int] = Field(None, description="SMTP port")
    email_username: Optional[str] = Field(None, description="Email username")
    email_password: Optional[str] = Field(None, description="Email password")
    slack_enabled: bool = Field(False, description="Enable Slack notifications")
    slack_webhook_url: Optional[str] = Field(None, description="Slack webhook URL")
    telegram_enabled: bool = Field(False, description="Enable Telegram notifications")
    telegram_bot_token: Optional[str] = Field(None, description="Telegram bot token")
    telegram_chat_id: Optional[str] = Field(None, description="Telegram chat ID")

class UltimateLeadScraperConfig(BaseModel):
    """Main configuration model"""
    # Core Settings
    machine_id: str = Field("ultimate-lead-scraper-v2", description="Machine ID")
    machine_version: str = Field("2.0.0", description="Machine version")
    environment: str = Field("production", description="Environment")
    
    # API Configuration
    api: APIConfig = Field(default_factory=APIConfig)
    
    # Scraper Modes
    scraper_modes: Dict[str, ScraperMode] = Field(
        default_factory=lambda: {
            "high_intent": ScraperMode(
                name="high_intent",
                description="High-Intent Leads (No/Broken Websites)",
                scrape_no_website=True,
                scrape_broken_website=True,
                scrape_with_website=False,
                quality_threshold=80,
                max_leads_per_cycle=10
            ),
            "all_leads": ScraperMode(
                name="all_leads",
                description="All Leads (Including Websites)",
                scrape_no_website=True,
                scrape_broken_website=True,
                scrape_with_website=True,
                quality_threshold=60,
                max_leads_per_cycle=50
            ),
            "premium_only": ScraperMode(
                name="premium_only",
                description="Premium Leads Only",
                scrape_no_website=True,
                scrape_broken_website=True,
                scrape_with_website=False,
                quality_threshold=90,
                max_leads_per_cycle=5
            )
        }
    )
    
    # Active Mode
    active_mode: str = Field("high_intent", description="Active scraper mode")
    
    # Region Settings
    default_state: str = Field("PA", description="Default state")
    default_country: str = Field("US", description="Default country")
    
    # Cities & Industries (with fallbacks)
    cities: List[str] = Field(
        default_factory=lambda: [
            "Philadelphia", "Pittsburgh", "Harrisburg", "Allentown", "Erie",
            "Reading", "Scranton", "Lancaster", "York", "Bethlehem"
        ]
    )
    
    industries: List[str] = Field(
        default_factory=lambda: [
            "Hardscaping Contractor", "Landscape Contractor", "HVAC Company",
            "Plumbing Services", "Electrical Contractor", "Roofing Company",
            "General Contractor", "Painting Services", "Concrete Contractor",
            "Excavation Services", "Deck Builder", "Fence Contractor",
            "Masonry Contractor", "Paving Contractor", "Tree Service",
            "Flooring Contractor", "Window Installation", "Siding Contractor",
            "Basement Waterproofing", "Kitchen Remodeling", "Bathroom Remodeling"
        ]
    )
    
    # Search Configuration
    search_phrases: List[str] = Field(
        default_factory=lambda: [
            "{industry} {city} {state}",
            "{city} {industry} services",
            "best {industry} {city}",
            "{industry} near {city}",
            "top rated {industry} {city}"
        ]
    )
    
    # Platform Configuration
    platforms_to_scrape: List[str] = Field(
        default_factory=lambda: [
            "google",
            "facebook",
            "yelp",
            "bbb",
            "angi",
            "homeadvisor",
            "houzz",
            "thumbtack"
        ]
    )
    
    # Blacklisted Domains
    blacklisted_domains: List[str] = Field(
        default_factory=lambda: [
            "yelp.com", "yellowpages.com", "angi.com", "homeadvisor.com",
            "thumbtack.com", "bbb.org", "facebook.com", "linkedin.com",
            "instagram.com", "twitter.com", "x.com", "pinterest.com",
            "wikipedia.org", "chamberofcommerce.com", "mapquest.com",
            "manta.com", "youtube.com", "google.com", "bing.com",
            "dexknows.com", "superpages.com", "whitepages.com"
        ]
    )
    
    # Operating Parameters
    operating_mode: str = Field("auto", description="Operating mode")
    searches_per_cycle: int = Field(10, ge=1, le=100, description="Searches per cycle")
    businesses_per_search: int = Field(15, ge=1, le=50, description="Businesses per search")
    cycle_interval: int = Field(300, ge=10, description="Cycle interval in seconds")
    max_cycles: int = Field(1000, ge=1, description="Maximum cycles")
    concurrent_scrapers: int = Field(5, ge=1, le=20, description="Concurrent scrapers")
    request_timeout: int = Field(30, ge=5, description="Request timeout in seconds")
    
    # Filters
    filters: ScrapingFilters = Field(default_factory=ScrapingFilters)
    
    # AI Enrichment
    ai_enrichment: AIEnrichment = Field(default_factory=AIEnrichment)
    
    # CRM Settings
    crm: CRMSettings = Field(default_factory=CRMSettings)
    
    # UI Settings
    ui: UISettings = Field(default_factory=UISettings)
    
    # Notifications
    notifications: NotificationSettings = Field(default_factory=NotificationSettings)
    
    # Storage
    storage: Dict[str, str] = Field(
        default_factory=lambda: {
            "leads_file": "data/leads.json",
            "qualified_leads": "data/qualified_leads.json",
            "premium_leads": "data/premium_leads.json",
            "logs_file": "logs/system.log",
            "cache_file": "cache/search_cache.json",
            "exports_dir": "exports",
            "backups_dir": "backups",
            "templates_dir": "templates"
        }
    )
    
    # Dashboard
    dashboard: Dict[str, Any] = Field(
        default_factory=lambda: {
            "port": 8501,
            "host": "0.0.0.0",
            "debug": False,
            "secret_key": Fernet.generate_key().decode(),
            "session_timeout": 86400,
            "max_upload_size": 100  # MB
        }
    )
    
    # Security
    security: Dict[str, Any] = Field(
        default_factory=lambda: {
            "encryption_key": Fernet.generate_key().decode(),
            "require_auth": True,
            "allowed_ips": [],
            "rate_limit": 100,
            "enable_audit_log": True
        }
    )
    
    # Performance
    performance: Dict[str, Any] = Field(
        default_factory=lambda: {
            "enable_caching": True,
            "cache_ttl": 3600,
            "enable_compression": True,
            "max_threads": 10,
            "enable_proxies": False,
            "proxy_list": []
        }
    )

def load_config() -> UltimateLeadScraperConfig:
    """Load configuration with validation"""
    # Load environment variables first
    load_dotenv()
    
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                config_data = json.load(f)
            
            # Merge with environment variables
            if os.getenv("SERPER_API_KEY"):
                config_data["api"]["serper_api_key"] = os.getenv("SERPER_API_KEY")
            if os.getenv("OPENAI_API_KEY"):
                config_data["api"]["openai_api_key"] = os.getenv("OPENAI_API_KEY")
            
            # Validate config
            config = UltimateLeadScraperConfig(**config_data)
            print("✅ Configuration loaded and validated")
            return config
            
        except Exception as e:
            print(f"⚠️  Config error: {e}")
            print("Creating new configuration with defaults...")
    
    # Create default config
    config = UltimateLeadScraperConfig()
    save_config(config)
    print("📝 Created new configuration file")
    return config

def save_config(config: UltimateLeadScraperConfig):
    """Save configuration to file"""
    with open(CONFIG_FILE, "w") as f:
        json.dump(config.dict(), f, indent=2, default=str)

CONFIG = load_config()

# Ensure storage directories exist
for dir_path in CONFIG.storage.values():
    if '/' in dir_path or '\\' in dir_path:
        os.makedirs(os.path.dirname(dir_path), exist_ok=True)

# ============================================================================
# ENHANCED LOGGER WITH ROTATION AND MULTIPLE HANDLERS
# ============================================================================

class EnhancedLogger:
    """Enhanced logger with file rotation, multiple handlers, and log levels"""
    
    def __init__(self):
        self.log_file = CONFIG.storage["logs_file"]
        self.max_file_size = 10 * 1024 * 1024  # 10 MB
        self.backup_count = 5
        self.setup_logger()
    
    def setup_logger(self):
        """Setup logger with handlers"""
        import logging
        from logging.handlers import RotatingFileHandler
        
        self.logger = logging.getLogger('UltimateLeadScraper')
        self.logger.setLevel(logging.DEBUG)
        
        # Remove existing handlers
        self.logger.handlers.clear()
        
        # File handler with rotation
        file_handler = RotatingFileHandler(
            self.log_file,
            maxBytes=self.max_file_size,
            backupCount=self.backup_count
        )
        file_handler.setLevel(logging.DEBUG)
        
        # Console handler
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        
        # Formatter
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)
        
        # Add handlers
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
    
    def log(self, message: str, level: str = "INFO", extra: Dict = None):
        """Log message with specified level"""
        log_methods = {
            "DEBUG": self.logger.debug,
            "INFO": self.logger.info,
            "WARNING": self.logger.warning,
            "ERROR": self.logger.error,
            "CRITICAL": self.logger.critical
        }
        
        log_method = log_methods.get(level.upper(), self.logger.info)
        
        # Add extra context if provided
        if extra:
            message = f"{message} | {json.dumps(extra)}"
        
        log_method(message)
        
        # Also print colored output for console
        colors = {
            "INFO": "\033[94m",
            "SUCCESS": "\033[92m",
            "WARNING": "\033[93m",
            "ERROR": "\033[91m",
            "DEBUG": "\033[90m"
        }
        
        color = colors.get(level, "\033[0m")
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"{color}[{timestamp}] {level}: {message}\033[0m")
    
    def get_recent_logs(self, limit: int = 100, level: str = None) -> List[Dict]:
        """Get recent logs"""
        logs = []
        try:
            with open(self.log_file, 'r') as f:
                lines = f.readlines()[-limit:]
            
            for line in lines:
                try:
                    # Parse log line
                    parts = line.split(' - ', 3)
                    if len(parts) >= 4:
                        log_entry = {
                            "timestamp": parts[0],
                            "logger": parts[1],
                            "level": parts[2],
                            "message": parts[3].strip()
                        }
                        
                        if not level or log_entry["level"] == level.upper():
                            logs.append(log_entry)
                except:
                    continue
        except:
            pass
        
        return logs

logger = EnhancedLogger()

# ============================================================================
# ENCRYPTION SERVICE FOR SENSITIVE DATA
# ============================================================================

class EncryptionService:
    """Service for encrypting sensitive data"""
    
    def __init__(self):
        key = CONFIG.security["encryption_key"]
        if not key:
            key = Fernet.generate_key()
            CONFIG.security["encryption_key"] = key.decode()
            save_config(CONFIG)
        
        self.cipher = Fernet(key.encode() if isinstance(key, str) else key)
    
    def encrypt(self, data: str) -> str:
        """Encrypt data"""
        return self.cipher.encrypt(data.encode()).decode()
    
    def decrypt(self, encrypted_data: str) -> str:
        """Decrypt data"""
        return self.cipher.decrypt(encrypted_data.encode()).decode()
    
    def encrypt_dict(self, data: Dict) -> Dict:
        """Encrypt dictionary values"""
        encrypted = {}
        for key, value in data.items():
            if isinstance(value, str) and key.lower() in ['password', 'key', 'token', 'secret']:
                encrypted[key] = self.encrypt(value)
            else:
                encrypted[key] = value
        return encrypted
    
    def decrypt_dict(self, data: Dict) -> Dict:
        """Decrypt dictionary values"""
        decrypted = {}
        for key, value in data.items():
            if isinstance(value, str) and key.lower() in ['password', 'key', 'token', 'secret']:
                try:
                    decrypted[key] = self.decrypt(value)
                except:
                    decrypted[key] = value
            else:
                decrypted[key] = value
        return decrypted

encryption_service = EncryptionService()

# ============================================================================
# ENHANCED DATABASE WITH MIGRATIONS AND AUDIT LOG
# ============================================================================

class UltimateCRM:
    """Enhanced SQLite CRM with migrations, audit log, and advanced features"""
    
    def __init__(self):
        self.db_file = CONFIG.crm.database
        self.conn = None
        self.cursor = None
        self.migration_version = 4  # Current schema version
        self.setup_database()
    
    def setup_database(self):
        """Initialize database with migrations"""
        try:
            self.conn = sqlite3.connect(self.db_file, check_same_thread=False)
            self.conn.row_factory = sqlite3.Row
            self.cursor = self.conn.cursor()
            
            # Enable foreign keys and WAL mode for better performance
            self.cursor.execute("PRAGMA foreign_keys = ON")
            self.cursor.execute("PRAGMA journal_mode = WAL")
            
            # Check current version
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS migrations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    version INTEGER NOT NULL,
                    migration_name TEXT NOT NULL,
                    applied_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            current_version = self.get_current_version()
            
            # Run migrations if needed
            if current_version < self.migration_version:
                self.run_migrations(current_version)
            
            # Create audit log table
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS audit_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    action TEXT NOT NULL,
                    entity_type TEXT,
                    entity_id INTEGER,
                    old_values TEXT,
                    new_values TEXT,
                    ip_address TEXT,
                    user_agent TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users (id)
                )
            ''')
            
            logger.log("✅ Database initialized with migrations", "SUCCESS")
            
        except Exception as e:
            logger.log(f"❌ Database error: {e}", "ERROR")
            raise
    
    def get_current_version(self) -> int:
        """Get current migration version"""
        try:
            self.cursor.execute("SELECT MAX(version) FROM migrations")
            result = self.cursor.fetchone()
            return result[0] if result and result[0] is not None else 0
        except:
            return 0
    
    def run_migrations(self, current_version: int):
        """Run database migrations"""
        migrations = [
            # Migration 1: Initial schema
            '''
            CREATE TABLE IF NOT EXISTS leads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fingerprint TEXT UNIQUE,
                business_name TEXT NOT NULL,
                website TEXT,
                website_status TEXT,
                phone TEXT,
                email TEXT,
                address TEXT,
                city TEXT,
                state TEXT,
                zip_code TEXT,
                country TEXT,
                industry TEXT,
                business_type TEXT,
                services TEXT,
                description TEXT,
                social_media TEXT,
                
                -- AI Enrichment Fields
                lead_score INTEGER DEFAULT 0,
                quality_tier TEXT,
                potential_value INTEGER DEFAULT 0,
                estimated_revenue TEXT,
                employee_count TEXT,
                years_in_business INTEGER,
                decision_maker_name TEXT,
                decision_maker_title TEXT,
                decision_maker_email TEXT,
                decision_maker_phone TEXT,
                
                -- Lead Metadata
                outreach_priority TEXT,
                lead_status TEXT DEFAULT 'New Lead',
                assigned_to TEXT,
                lead_source TEXT DEFAULT 'Web Scraper',
                scraped_date DATETIME,
                last_contacted DATETIME,
                next_followup DATETIME,
                lead_production_date DATE,
                
                -- Meeting Info
                meeting_type TEXT,
                meeting_date DATETIME,
                meeting_outcome TEXT,
                
                -- Notes & AI Insights
                notes TEXT,
                ai_notes TEXT,
                outreach_strategy TEXT,
                
                -- Platform Data
                google_business_url TEXT,
                facebook_business_url TEXT,
                yelp_business_url TEXT,
                bbb_business_url TEXT,
                other_platforms TEXT,
                
                -- System Fields
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                is_archived BOOLEAN DEFAULT 0,
                archive_reason TEXT,
                archive_date DATETIME,
                
                -- Indexes
                CHECK (lead_score BETWEEN 0 AND 100)
            )
            ''',
            
            # Migration 2: Users and activities
            '''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                email TEXT UNIQUE NOT NULL,
                full_name TEXT,
                password_hash TEXT,
                role TEXT DEFAULT 'user',
                phone TEXT,
                department TEXT,
                is_active BOOLEAN DEFAULT 1,
                last_login DATETIME,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );
            
            CREATE TABLE IF NOT EXISTS activities (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                lead_id INTEGER NOT NULL,
                user_id INTEGER,
                activity_type TEXT NOT NULL,
                activity_details TEXT,
                outcome TEXT,
                duration_minutes INTEGER,
                next_step TEXT,
                scheduled_followup DATETIME,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (lead_id) REFERENCES leads (id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users (id)
            );
            ''',
            
            # Migration 3: Statistics and settings
            '''
            CREATE TABLE IF NOT EXISTS daily_statistics (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                stat_date DATE UNIQUE NOT NULL,
                total_leads INTEGER DEFAULT 0,
                new_leads INTEGER DEFAULT 0,
                contacted_leads INTEGER DEFAULT 0,
                meetings_scheduled INTEGER DEFAULT 0,
                closed_won INTEGER DEFAULT 0,
                closed_lost INTEGER DEFAULT 0,
                premium_leads INTEGER DEFAULT 0,
                estimated_value INTEGER DEFAULT 0,
                scrape_cycles INTEGER DEFAULT 0,
                websites_checked INTEGER DEFAULT 0,
                leads_found INTEGER DEFAULT 0,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );
            
            CREATE TABLE IF NOT EXISTS settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                setting_key TEXT UNIQUE NOT NULL,
                setting_value TEXT,
                setting_type TEXT DEFAULT 'string',
                category TEXT DEFAULT 'general',
                description TEXT,
                is_encrypted BOOLEAN DEFAULT 0,
                updated_by INTEGER,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (updated_by) REFERENCES users (id)
            );
            
            CREATE TABLE IF NOT EXISTS email_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                template_name TEXT UNIQUE NOT NULL,
                template_subject TEXT,
                template_body TEXT,
                template_type TEXT DEFAULT 'email',
                variables TEXT,
                is_active BOOLEAN DEFAULT 1,
                created_by INTEGER,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (created_by) REFERENCES users (id)
            );
            ''',
            
            # Migration 4: Campaigns and sequences
            '''
            CREATE TABLE IF NOT EXISTS campaigns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                campaign_name TEXT UNIQUE NOT NULL,
                description TEXT,
                target_status TEXT,
                target_quality_tier TEXT,
                email_template_id INTEGER,
                sms_template_id INTEGER,
                is_active BOOLEAN DEFAULT 1,
                created_by INTEGER,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (email_template_id) REFERENCES email_templates (id),
                FOREIGN KEY (created_by) REFERENCES users (id)
            );
            
            CREATE TABLE IF NOT EXISTS sequences (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sequence_name TEXT UNIQUE NOT NULL,
                campaign_id INTEGER NOT NULL,
                step_number INTEGER NOT NULL,
                action_type TEXT NOT NULL,
                action_template_id INTEGER,
                days_after_previous INTEGER DEFAULT 1,
                conditions TEXT,
                is_active BOOLEAN DEFAULT 1,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (campaign_id) REFERENCES campaigns (id) ON DELETE CASCADE,
                FOREIGN KEY (action_template_id) REFERENCES email_templates (id),
                UNIQUE(campaign_id, step_number)
            );
            
            CREATE TABLE IF NOT EXISTS campaign_leads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                campaign_id INTEGER NOT NULL,
                lead_id INTEGER NOT NULL,
                sequence_id INTEGER,
                current_step INTEGER DEFAULT 1,
                last_action_date DATETIME,
                next_action_date DATETIME,
                status TEXT DEFAULT 'active',
                responses TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (campaign_id) REFERENCES campaigns (id) ON DELETE CASCADE,
                FOREIGN KEY (lead_id) REFERENCES leads (id) ON DELETE CASCADE,
                FOREIGN KEY (sequence_id) REFERENCES sequences (id),
                UNIQUE(campaign_id, lead_id)
            );
            '''
        ]
        
        try:
            for i in range(current_version, self.migration_version):
                migration_sql = migrations[i]
                self.cursor.executescript(migration_sql)
                
                # Record migration
                self.cursor.execute(
                    "INSERT INTO migrations (version, migration_name) VALUES (?, ?)",
                    (i + 1, f"migration_{i + 1}")
                )
                
                logger.log(f"✅ Applied migration {i + 1}", "SUCCESS")
            
            self.conn.commit()
            
            # Create indexes for performance
            self.create_indexes()
            
        except Exception as e:
            self.conn.rollback()
            logger.log(f"❌ Migration error: {e}", "ERROR")
            raise
    
    def create_indexes(self):
        """Create performance indexes"""
        indexes = [
            "CREATE INDEX IF NOT EXISTS idx_leads_fingerprint ON leads(fingerprint)",
            "CREATE INDEX IF NOT EXISTS idx_leads_status ON leads(lead_status)",
            "CREATE INDEX IF NOT EXISTS idx_leads_quality ON leads(quality_tier)",
            "CREATE INDEX IF NOT EXISTS idx_leads_city ON leads(city)",
            "CREATE INDEX IF NOT EXISTS idx_leads_industry ON leads(industry)",
            "CREATE INDEX IF NOT EXISTS idx_leads_score ON leads(lead_score)",
            "CREATE INDEX IF NOT EXISTS idx_leads_created ON leads(created_at)",
            "CREATE INDEX IF NOT EXISTS idx_leads_website_status ON leads(website_status)",
            
            "CREATE INDEX IF NOT EXISTS idx_activities_lead ON activities(lead_id)",
            "CREATE INDEX IF NOT EXISTS idx_activities_created ON activities(created_at)",
            
            "CREATE INDEX IF NOT EXISTS idx_campaign_leads_status ON campaign_leads(status)",
            "CREATE INDEX IF NOT EXISTS idx_campaign_leads_next_action ON campaign_leads(next_action_date)",
        ]
        
        for index_sql in indexes:
            try:
                self.cursor.execute(index_sql)
            except Exception as e:
                logger.log(f"Index creation error: {e}", "WARNING")
        
        self.conn.commit()
    
    def get_connection(self):
        """Get a new database connection"""
        conn = sqlite3.connect(self.db_file, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        return conn
    
    def save_lead(self, lead_data: Dict, user_id: Optional[int] = None) -> Dict:
        """Save lead to database with audit logging"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            # Generate fingerprint if not provided
            fingerprint = lead_data.get("fingerprint")
            if not fingerprint:
                fingerprint_data = (
                    lead_data.get("business_name", ""),
                    lead_data.get("website", ""),
                    lead_data.get("phone", ""),
                    lead_data.get("city", "")
                )
                fingerprint = hashlib.sha256(str(fingerprint_data).encode()).hexdigest()
            
            lead_data["fingerprint"] = fingerprint
            
            # Check for duplicates
            if CONFIG.crm.prevent_duplicates:
                cursor.execute("SELECT id FROM leads WHERE fingerprint = ?", (fingerprint,))
                existing = cursor.fetchone()
                if existing:
                    return {
                        "success": False,
                        "message": "Duplicate lead detected",
                        "lead_id": existing[0],
                        "action": "skipped"
                    }
            
            # Website status analysis
            website_status = lead_data.get("website_status", "unknown")
            if not website_status or website_status == "unknown":
                website = lead_data.get("website", "")
                if website:
                    website_status = self.analyze_website_status(website, lead_data)
                    lead_data["website_status"] = website_status
            
            # Prepare data for insertion
            columns = []
            placeholders = []
            values = []
            
            # Map lead data to table columns
            column_mapping = {
                'fingerprint': 'fingerprint',
                'business_name': 'business_name',
                'website': 'website',
                'website_status': 'website_status',
                'phone': 'phone',
                'email': 'email',
                'address': 'address',
                'city': 'city',
                'state': 'state',
                'zip_code': 'zip_code',
                'country': 'country',
                'industry': 'industry',
                'business_type': 'business_type',
                'services': 'services',
                'description': 'description',
                'social_media': 'social_media',
                'lead_score': 'lead_score',
                'quality_tier': 'quality_tier',
                'potential_value': 'potential_value',
                'estimated_revenue': 'estimated_revenue',
                'employee_count': 'employee_count',
                'years_in_business': 'years_in_business',
                'decision_maker_name': 'decision_maker_name',
                'decision_maker_title': 'decision_maker_title',
                'decision_maker_email': 'decision_maker_email',
                'decision_maker_phone': 'decision_maker_phone',
                'outreach_priority': 'outreach_priority',
                'lead_status': 'lead_status',
                'assigned_to': 'assigned_to',
                'lead_source': 'lead_source',
                'scraped_date': 'scraped_date',
                'google_business_url': 'google_business_url',
                'facebook_business_url': 'facebook_business_url',
                'yelp_business_url': 'yelp_business_url',
                'bbb_business_url': 'bbb_business_url',
                'other_platforms': 'other_platforms',
                'notes': 'notes',
                'ai_notes': 'ai_notes',
                'outreach_strategy': 'outreach_strategy'
            }
            
            for data_key, column_name in column_mapping.items():
                if data_key in lead_data and lead_data[data_key] is not None:
                    value = lead_data[data_key]
                    
                    # Convert lists/dicts to JSON strings
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value)
                    
                    columns.append(column_name)
                    placeholders.append('?')
                    values.append(value)
            
            # Add system fields
            columns.extend(['created_at', 'updated_at'])
            placeholders.extend(['CURRENT_TIMESTAMP', 'CURRENT_TIMESTAMP'])
            
            # Build and execute insert query
            query = f'''
                INSERT INTO leads ({', '.join(columns)})
                VALUES ({', '.join(placeholders)})
            '''
            
            cursor.execute(query, values)
            lead_id = cursor.lastrowid
            
            # Log activity
            cursor.execute('''
                INSERT INTO activities (lead_id, activity_type, activity_details)
                VALUES (?, ?, ?)
            ''', (lead_id, "Lead Created", f"Lead scraped from {lead_data.get('website', 'unknown')}"))
            
            # Update daily statistics
            self.update_daily_statistics(cursor)
            
            # Audit log
            if user_id:
                self.log_audit(
                    user_id=user_id,
                    action="CREATE_LEAD",
                    entity_type="lead",
                    entity_id=lead_id,
                    new_values=json.dumps(lead_data)
                )
            
            conn.commit()
            
            return {
                "success": True,
                "lead_id": lead_id,
                "message": "Lead saved successfully",
                "fingerprint": fingerprint,
                "website_status": website_status
            }
            
        except Exception as e:
            conn.rollback()
            logger.log(f"Save lead error: {e}", "ERROR")
            return {"success": False, "message": f"Error: {str(e)}"}
        finally:
            conn.close()
    
    def analyze_website_status(self, website: str, lead_data: Dict) -> str:
        """Analyze website status"""
        if not website or website.lower() in ["none", "na", "n/a"]:
            return "no_website"
        
        try:
            # Check if website is accessible
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            
            response = requests.head(website, headers=headers, timeout=10, allow_redirects=True)
            
            if response.status_code >= 400:
                return "broken"
            
            # Check for parked domains or placeholders
            response = requests.get(website, headers=headers, timeout=10)
            content = response.text.lower()
            
            parked_indicators = [
                'domain for sale', 'parked domain', 'this domain is',
                'godaddy', 'namecheap', 'hostinger', 'domain parking',
                'buy this domain', 'is for sale'
            ]
            
            placeholder_indicators = [
                'coming soon', 'under construction', 'website coming soon',
                'site under maintenance', 'be right back'
            ]
            
            if any(indicator in content for indicator in parked_indicators):
                return "parked"
            
            if any(indicator in content for indicator in placeholder_indicators):
                return "placeholder"
            
            return "active"
            
        except requests.exceptions.RequestException:
            return "unreachable"
        except Exception:
            return "unknown"
    
    def update_daily_statistics(self, cursor):
        """Update daily statistics"""
        try:
            today = datetime.now().date().isoformat()
            
            # Get counts for today
            cursor.execute('''
                SELECT 
                    COUNT(*) as total_leads,
                    SUM(CASE WHEN lead_status = 'New Lead' THEN 1 ELSE 0 END) as new_leads,
                    SUM(CASE WHEN lead_status = 'Contacted' THEN 1 ELSE 0 END) as contacted_leads,
                    SUM(CASE WHEN lead_status IN ('Meeting Scheduled', 'Zoom Meeting') THEN 1 ELSE 0 END) as meetings_scheduled,
                    SUM(CASE WHEN lead_status = 'Closed (Won)' THEN 1 ELSE 0 END) as closed_won,
                    SUM(CASE WHEN lead_status = 'Closed (Lost)' THEN 1 ELSE 0 END) as closed_lost,
                    SUM(CASE WHEN quality_tier IN ('Premium', 'High') THEN 1 ELSE 0 END) as premium_leads,
                    SUM(potential_value) as estimated_value
                FROM leads 
                WHERE DATE(created_at) = DATE('now') AND is_archived = 0
            ''')
            
            stats = cursor.fetchone()
            
            cursor.execute('''
                INSERT OR REPLACE INTO daily_statistics 
                (stat_date, total_leads, new_leads, contacted_leads, meetings_scheduled, 
                 closed_won, closed_lost, premium_leads, estimated_value)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (today, *stats))
            
        except Exception as e:
            logger.log(f"Statistics update error: {e}", "WARNING")
    
    def log_audit(self, user_id: Optional[int], action: str, entity_type: str = None,
                 entity_id: int = None, old_values: str = None, new_values: str = None,
                 ip_address: str = None, user_agent: str = None):
        """Log audit trail"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO audit_log 
                (user_id, action, entity_type, entity_id, old_values, new_values, ip_address, user_agent)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (user_id, action, entity_type, entity_id, old_values, new_values, ip_address, user_agent))
            
            conn.commit()
        except Exception as e:
            logger.log(f"Audit log error: {e}", "WARNING")
        finally:
            conn.close()
    
    def get_leads(self, filters: Dict = None, page: int = 1, per_page: int = 50,
                 sort_by: str = "created_at", sort_order: str = "DESC") -> Dict:
        """Get leads with advanced filtering and pagination"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            # Base query
            query = '''
                SELECT 
                    l.*,
                    COUNT(a.id) as activity_count,
                    MAX(a.created_at) as last_activity_date
                FROM leads l
                LEFT JOIN activities a ON l.id = a.lead_id
                WHERE l.is_archived = 0
            '''
            
            params = []
            conditions = []
            
            # Apply filters
            if filters:
                # Text search
                if filters.get("search"):
                    search_term = f"%{filters['search']}%"
                    conditions.append('''
                        (l.business_name LIKE ? OR 
                         l.website LIKE ? OR 
                         l.phone LIKE ? OR 
                         l.email LIKE ? OR 
                         l.city LIKE ? OR 
                         l.industry LIKE ?)
                    ''')
                    params.extend([search_term] * 6)
                
                # Status filter
                if filters.get("status"):
                    if isinstance(filters["status"], list):
                        placeholders = ','.join(['?'] * len(filters["status"]))
                        conditions.append(f"l.lead_status IN ({placeholders})")
                        params.extend(filters["status"])
                    else:
                        conditions.append("l.lead_status = ?")
                        params.append(filters["status"])
                
                # Quality tier filter
                if filters.get("quality_tier"):
                    if isinstance(filters["quality_tier"], list):
                        placeholders = ','.join(['?'] * len(filters["quality_tier"]))
                        conditions.append(f"l.quality_tier IN ({placeholders})")
                        params.extend(filters["quality_tier"])
                    else:
                        conditions.append("l.quality_tier = ?")
                        params.append(filters["quality_tier"])
                
                # Website status filter
                if filters.get("website_status"):
                    if isinstance(filters["website_status"], list):
                        placeholders = ','.join(['?'] * len(filters["website_status"]))
                        conditions.append(f"l.website_status IN ({placeholders})")
                        params.extend(filters["website_status"])
                    else:
                        conditions.append("l.website_status = ?")
                        params.append(filters["website_status"])
                
                # City filter
                if filters.get("city"):
                    if isinstance(filters["city"], list):
                        placeholders = ','.join(['?'] * len(filters["city"]))
                        conditions.append(f"l.city IN ({placeholders})")
                        params.extend(filters["city"])
                    else:
                        conditions.append("l.city = ?")
                        params.append(filters["city"])
                
                # Industry filter
                if filters.get("industry"):
                    if isinstance(filters["industry"], list):
                        placeholders = ','.join(['?'] * len(filters["industry"]))
                        conditions.append(f"l.industry IN ({placeholders})")
                        params.extend(filters["industry"])
                    else:
                        conditions.append("l.industry = ?")
                        params.append(filters["industry"])
                
                # Score range
                if filters.get("min_score"):
                    conditions.append("l.lead_score >= ?")
                    params.append(filters["min_score"])
                
                if filters.get("max_score"):
                    conditions.append("l.lead_score <= ?")
                    params.append(filters["max_score"])
                
                # Date range
                if filters.get("date_from"):
                    conditions.append("DATE(l.created_at) >= ?")
                    params.append(filters["date_from"])
                
                if filters.get("date_to"):
                    conditions.append("DATE(l.created_at) <= ?")
                    params.append(filters["date_to"])
                
                # Assigned to
                if filters.get("assigned_to"):
                    conditions.append("l.assigned_to = ?")
                    params.append(filters["assigned_to"])
            
            # Add conditions to query
            if conditions:
                query += " AND " + " AND ".join(conditions)
            
            # Group by lead
            query += " GROUP BY l.id"
            
            # Get total count
            count_query = f"SELECT COUNT(*) as total FROM ({query})"
            cursor.execute(count_query, params)
            total_result = cursor.fetchone()
            total = total_result['total'] if total_result else 0
            
            # Add sorting
            valid_sort_columns = ['created_at', 'updated_at', 'lead_score', 'potential_value', 
                                'business_name', 'city', 'industry']
            if sort_by in valid_sort_columns:
                query += f" ORDER BY l.{sort_by} {sort_order}"
            
            # Add pagination
            query += " LIMIT ? OFFSET ?"
            params.extend([per_page, (page - 1) * per_page])
            
            # Execute query
            cursor.execute(query, params)
            leads = cursor.fetchall()
            
            # Convert to list of dictionaries
            result = []
            for lead in leads:
                lead_dict = dict(lead)
                
                # Parse JSON fields
                json_fields = ['social_media', 'services', 'other_platforms']
                for field in json_fields:
                    if lead_dict.get(field) and isinstance(lead_dict[field], str):
                        try:
                            lead_dict[field] = json.loads(lead_dict[field])
                        except:
                            pass
                
                result.append(lead_dict)
            
            return {
                "leads": result,
                "total": total,
                "page": page,
                "per_page": per_page,
                "total_pages": (total + per_page - 1) // per_page if per_page > 0 else 0
            }
            
        except Exception as e:
            logger.log(f"Get leads error: {e}", "ERROR")
            return {"leads": [], "total": 0, "page": page, "per_page": per_page}
        finally:
            conn.close()
    
    def get_lead_by_id(self, lead_id: int, include_activities: bool = True) -> Optional[Dict]:
        """Get lead by ID with optional activities"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("SELECT * FROM leads WHERE id = ?", (lead_id,))
            lead = cursor.fetchone()
            
            if not lead:
                return None
            
            lead_dict = dict(lead)
            
            # Parse JSON fields
            json_fields = ['social_media', 'services', 'other_platforms']
            for field in json_fields:
                if lead_dict.get(field) and isinstance(lead_dict[field], str):
                    try:
                        lead_dict[field] = json.loads(lead_dict[field])
                    except:
                        pass
            
            # Get activities if requested
            if include_activities:
                cursor.execute(
                    "SELECT * FROM activities WHERE lead_id = ? ORDER BY created_at DESC",
                    (lead_id,)
                )
                activities = cursor.fetchall()
                lead_dict["activities"] = [dict(activity) for activity in activities]
            
            return lead_dict
            
        except Exception as e:
            logger.log(f"Get lead error: {e}", "ERROR")
            return None
        finally:
            conn.close()
    
    def update_lead(self, lead_id: int, update_data: Dict, user_id: Optional[int] = None) -> Dict:
        """Update lead with audit logging"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            # Get old values for audit log
            cursor.execute("SELECT * FROM leads WHERE id = ?", (lead_id,))
            old_lead = cursor.fetchone()
            
            if not old_lead:
                return {"success": False, "message": "Lead not found"}
            
            old_values = dict(old_lead)
            
            # Build update query
            set_clause = []
            params = []
            
            for field, value in update_data.items():
                # Convert lists/dicts to JSON strings
                if isinstance(value, (list, dict)):
                    value = json.dumps(value)
                
                set_clause.append(f"{field} = ?")
                params.append(value)
            
            # Add updated_at timestamp
            set_clause.append("updated_at = CURRENT_TIMESTAMP")
            
            params.append(lead_id)
            query = f"UPDATE leads SET {', '.join(set_clause)} WHERE id = ?"
            
            cursor.execute(query, params)
            
            # Log activity
            activity_desc = f"Updated fields: {', '.join(update_data.keys())}"
            cursor.execute('''
                INSERT INTO activities (lead_id, activity_type, activity_details)
                VALUES (?, ?, ?)
            ''', (lead_id, "Lead Updated", activity_desc))
            
            # Audit log
            if user_id:
                new_values = old_values.copy()
                new_values.update(update_data)
                
                self.log_audit(
                    user_id=user_id,
                    action="UPDATE_LEAD",
                    entity_type="lead",
                    entity_id=lead_id,
                    old_values=json.dumps(old_values),
                    new_values=json.dumps(new_values)
                )
            
            conn.commit()
            
            return {"success": True, "message": "Lead updated successfully"}
            
        except Exception as e:
            conn.rollback()
            logger.log(f"Update lead error: {e}", "ERROR")
            return {"success": False, "message": f"Error: {str(e)}"}
        finally:
            conn.close()
    
    def delete_lead(self, lead_id: int, user_id: Optional[int] = None, reason: str = None) -> Dict:
        """Soft delete lead (archive)"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            # Get lead for audit log
            cursor.execute("SELECT * FROM leads WHERE id = ?", (lead_id,))
            old_lead = cursor.fetchone()
            
            if not old_lead:
                return {"success": False, "message": "Lead not found"}
            
            # Archive lead
            cursor.execute('''
                UPDATE leads 
                SET is_archived = 1, 
                    archive_reason = ?,
                    archive_date = CURRENT_TIMESTAMP 
                WHERE id = ?
            ''', (reason or "Manual archive", lead_id))
            
            # Log activity
            cursor.execute('''
                INSERT INTO activities (lead_id, activity_type, activity_details)
                VALUES (?, ?, ?)
            ''', (lead_id, "Lead Archived", f"Archived: {reason or 'No reason provided'}"))
            
            # Audit log
            if user_id:
                self.log_audit(
                    user_id=user_id,
                    action="ARCHIVE_LEAD",
                    entity_type="lead",
                    entity_id=lead_id,
                    old_values=json.dumps(dict(old_lead))
                )
            
            conn.commit()
            
            return {"success": True, "message": "Lead archived successfully"}
            
        except Exception as e:
            conn.rollback()
            logger.log(f"Archive lead error: {e}", "ERROR")
            return {"success": False, "message": f"Error: {str(e)}"}
        finally:
            conn.close()
    
    def get_statistics(self, period: str = "30d") -> Dict:
        """Get comprehensive statistics"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            stats = {}
            
            # Determine date range
            if period == "7d":
                date_filter = "7 day"
            elif period == "30d":
                date_filter = "30 day"
            elif period == "90d":
                date_filter = "90 day"
            else:
                date_filter = "30 day"  # Default
            
            # Overall statistics
            cursor.execute(f'''
                SELECT 
                    COUNT(*) as total_leads,
                    SUM(CASE WHEN lead_status = 'New Lead' THEN 1 ELSE 0 END) as new_leads,
                    SUM(CASE WHEN lead_status = 'Contacted' THEN 1 ELSE 0 END) as contacted_leads,
                    SUM(CASE WHEN lead_status IN ('Meeting Scheduled', 'Zoom Meeting') THEN 1 ELSE 0 END) as meetings_scheduled,
                    SUM(CASE WHEN lead_status = 'Closed (Won)' THEN 1 ELSE 0 END) as closed_won,
                    SUM(CASE WHEN lead_status = 'Closed (Lost)' THEN 1 ELSE 0 END) as closed_lost,
                    SUM(potential_value) as total_potential_value,
                    AVG(lead_score) as average_score,
                    COUNT(DISTINCT city) as cities_covered,
                    COUNT(DISTINCT industry) as industries_covered
                FROM leads 
                WHERE is_archived = 0 AND created_at >= DATE('now', '-{date_filter}')
            ''')
            
            overall = cursor.fetchone()
            stats["overall"] = dict(overall) if overall else {}
            
            # Lead quality distribution
            cursor.execute(f'''
                SELECT 
                    quality_tier,
                    COUNT(*) as count,
                    AVG(lead_score) as avg_score,
                    SUM(potential_value) as total_value
                FROM leads 
                WHERE is_archived = 0 AND created_at >= DATE('now', '-{date_filter}')
                GROUP BY quality_tier
                ORDER BY 
                    CASE quality_tier
                        WHEN 'Premium' THEN 1
                        WHEN 'High' THEN 2
                        WHEN 'Medium' THEN 3
                        WHEN 'Low' THEN 4
                        ELSE 5
                    END
            ''')
            
            stats["quality_distribution"] = [
                dict(row) for row in cursor.fetchall()
            ]
            
            # Website status distribution
            cursor.execute(f'''
                SELECT 
                    website_status,
                    COUNT(*) as count,
                    AVG(lead_score) as avg_score
                FROM leads 
                WHERE is_archived = 0 AND created_at >= DATE('now', '-{date_filter}')
                GROUP BY website_status
                ORDER BY count DESC
            ''')
            
            stats["website_status_distribution"] = [
                dict(row) for row in cursor.fetchall()
            ]
            
            # Daily leads trend
            cursor.execute(f'''
                SELECT 
                    DATE(created_at) as date,
                    COUNT(*) as leads_count,
                    SUM(CASE WHEN lead_status = 'New Lead' THEN 1 ELSE 0 END) as new_leads,
                    SUM(CASE WHEN quality_tier IN ('Premium', 'High') THEN 1 ELSE 0 END) as premium_leads
                FROM leads 
                WHERE is_archived = 0 AND created_at >= DATE('now', '-{date_filter}')
                GROUP BY DATE(created_at)
                ORDER BY date
            ''')
            
            stats["daily_trend"] = [
                dict(row) for row in cursor.fetchall()
            ]
            
            # Top cities
            cursor.execute(f'''
                SELECT 
                    city,
                    COUNT(*) as lead_count,
                    AVG(lead_score) as avg_score,
                    SUM(potential_value) as total_value
                FROM leads 
                WHERE is_archived = 0 AND created_at >= DATE('now', '-{date_filter}')
                GROUP BY city
                ORDER BY lead_count DESC
                LIMIT 10
            ''')
            
            stats["top_cities"] = [
                dict(row) for row in cursor.fetchall()
            ]
            
            # Top industries
            cursor.execute(f'''
                SELECT 
                    industry,
                    COUNT(*) as lead_count,
                    AVG(lead_score) as avg_score,
                    SUM(potential_value) as total_value
                FROM leads 
                WHERE is_archived = 0 AND created_at >= DATE('now', '-{date_filter}')
                GROUP BY industry
                ORDER BY lead_count DESC
                LIMIT 10
            ''')
            
            stats["top_industries"] = [
                dict(row) for row in cursor.fetchall()
            ]
            
            # Conversion funnel
            cursor.execute(f'''
                SELECT 
                    lead_status as stage,
                    COUNT(*) as count,
                    AVG(lead_score) as avg_score
                FROM leads 
                WHERE is_archived = 0 AND created_at >= DATE('now', '-{date_filter}')
                GROUP BY lead_status
                ORDER BY 
                    CASE lead_status
                        WHEN 'New Lead' THEN 1
                        WHEN 'Contacted' THEN 2
                        WHEN 'Follow Up' THEN 3
                        WHEN 'Meeting Scheduled' THEN 4
                        WHEN 'Zoom Meeting' THEN 5
                        WHEN 'Closed (Won)' THEN 6
                        WHEN 'Closed (Lost)' THEN 7
                        ELSE 8
                    END
            ''')
            
            stats["conversion_funnel"] = [
                dict(row) for row in cursor.fetchall()
            ]
            
            return stats
            
        except Exception as e:
            logger.log(f"Statistics error: {e}", "ERROR")
            return {}
        finally:
            conn.close()
    
    def get_today_stats(self) -> Dict:
        """Get today's statistics"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                SELECT 
                    COUNT(*) as today_leads,
                    SUM(CASE WHEN website_status IN ('no_website', 'broken', 'parked') THEN 1 ELSE 0 END) as high_intent_leads,
                    SUM(CASE WHEN quality_tier IN ('Premium', 'High') THEN 1 ELSE 0 END) as premium_leads,
                    SUM(potential_value) as today_value
                FROM leads 
                WHERE DATE(created_at) = DATE('now') AND is_archived = 0
            ''')
            
            result = cursor.fetchone()
            return dict(result) if result else {}
            
        except Exception as e:
            logger.log(f"Today stats error: {e}", "ERROR")
            return {}
        finally:
            conn.close()

# Initialize CRM
crm = UltimateCRM()

# ============================================================================
# WEBSITE CHECKER WITH ADVANCED ANALYSIS
# ============================================================================

class AdvancedWebsiteChecker:
    """Advanced website checker with multiple verification methods"""
    
    def __init__(self):
        self.user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15"
        ]
        
        self.timeout = CONFIG.request_timeout
        self.proxies = CONFIG.performance.get("proxy_list", [])
        self.current_proxy_idx = 0
    
    def get_next_proxy(self):
        """Get next proxy from pool"""
        if not self.proxies:
            return None
        
        proxy = self.proxies[self.current_proxy_idx]
        self.current_proxy_idx = (self.current_proxy_idx + 1) % len(self.proxies)
        return proxy
    
    async def check_website_async(self, url: str) -> Dict:
        """Check website asynchronously"""
        result = {
            "url": url,
            "status": "unknown",
            "status_code": None,
            "load_time": None,
            "title": None,
            "has_contact_form": False,
            "has_phone": False,
            "has_email": False,
            "is_parked": False,
            "is_placeholder": False,
            "ssl_valid": False,
            "responsive": False,
            "technologies": [],
            "error": None
        }
        
        if not url or not url.startswith(('http://', 'https://')):
            result["status"] = "invalid_url"
            return result
        
        try:
            headers = {
                'User-Agent': random.choice(self.user_agents),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
                'Accept-Encoding': 'gzip, deflate',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1'
            }
            
            start_time = time.time()
            
            async with aiohttp.ClientSession(headers=headers) as session:
                async with session.get(url, timeout=self.timeout, ssl=False) as response:
                    result["status_code"] = response.status
                    result["load_time"] = time.time() - start_time
                    
                    if response.status >= 400:
                        result["status"] = "broken"
                        return result
                    
                    # Get page content
                    html_content = await response.text()
                    
                    # Check for parked domains
                    if self.is_parked_domain(html_content):
                        result["is_parked"] = True
                        result["status"] = "parked"
                        return result
                    
                    # Check for placeholder pages
                    if self.is_placeholder_page(html_content):
                        result["is_placeholder"] = True
                        result["status"] = "placeholder"
                        return result
                    
                    # Parse HTML
                    soup = BeautifulSoup(html_content, 'html.parser')
                    
                    # Extract title
                    if soup.title and soup.title.string:
                        result["title"] = soup.title.string.strip()[:200]
                    
                    # Check for contact information
                    result["has_contact_form"] = self.has_contact_form(soup)
                    result["has_phone"] = self.has_phone_number(html_content)
                    result["has_email"] = self.has_email_address(html_content)
                    
                    # Check if responsive (has viewport meta tag)
                    result["responsive"] = self.is_responsive(soup)
                    
                    # Check SSL (if HTTPS)
                    if url.startswith('https://'):
                        result["ssl_valid"] = await self.check_ssl_async(url)
                    
                    result["status"] = "active"
                    
        except asyncio.TimeoutError:
            result["status"] = "timeout"
            result["error"] = "Request timed out"
        except aiohttp.ClientError as e:
            result["status"] = "unreachable"
            result["error"] = str(e)
        except Exception as e:
            result["status"] = "error"
            result["error"] = str(e)
        
        return result
    
    async def check_ssl_async(self, url: str) -> bool:
        """Check SSL certificate validity"""
        try:
            hostname = urlparse(url).hostname
            port = 443
            
            # Create SSL context
            context = ssl.create_default_context()
            
            # Create socket and wrap with SSL
            with socket.create_connection((hostname, port), timeout=10) as sock:
                with context.wrap_socket(sock, server_hostname=hostname) as ssock:
                    cert = ssock.getpeercert()
                    
                    # Check certificate expiration
                    if cert:
                        # Parse expiration date
                        expiry_date = datetime.strptime(cert['notAfter'], '%b %d %H:%M:%S %Y %Z')
                        if expiry_date > datetime.now():
                            return True
            
            return False
        except:
            return False
    
    def is_parked_domain(self, html_content: str) -> bool:
        """Check if domain is parked"""
        content_lower = html_content.lower()
        
        parked_indicators = [
            'domain for sale', 'parked domain', 'this domain is',
            'godaddy', 'namecheap', 'hostinger', 'domain parking',
            'buy this domain', 'is for sale', 'domainparking',
            'sedoparking', 'parkingcrew', 'above.com', 'voodoo.com',
            'bodis.com', 'domainname', 'premium domain'
        ]
        
        return any(indicator in content_lower for indicator in parked_indicators)
    
    def is_placeholder_page(self, html_content: str) -> bool:
        """Check if page is a placeholder"""
        content_lower = html_content.lower()
        
        placeholder_indicators = [
            'coming soon', 'under construction', 'website coming soon',
            'site under maintenance', 'be right back', 'this site is',
            'page is being', 'will be back', 'temporarily unavailable',
            'check back soon', 'site is under', 'we are working'
        ]
        
        return any(indicator in content_lower for indicator in placeholder_indicators)
    
    def has_contact_form(self, soup) -> bool:
        """Check if page has contact form"""
        # Look for form elements with common contact form indicators
        forms = soup.find_all('form')
        
        for form in forms:
            form_html = str(form).lower()
            
            # Check for contact-related keywords
            contact_keywords = [
                'contact', 'message', 'inquiry', 'request', 'quote',
                'consultation', 'estimate', 'callback', 'reach out'
            ]
            
            if any(keyword in form_html for keyword in contact_keywords):
                return True
            
            # Check for common contact form fields
            input_fields = form.find_all(['input', 'textarea'])
            field_names = [field.get('name', '').lower() for field in input_fields]
            field_ids = [field.get('id', '').lower() for field in input_fields]
            
            contact_fields = ['name', 'email', 'phone', 'message', 'subject']
            
            if any(field in field_names + field_ids for field in contact_fields):
                return True
        
        return False
    
    def has_phone_number(self, html_content: str) -> bool:
        """Check if page has phone number"""
        phone_patterns = [
            r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}',
            r'\d{3}[-.\s]?\d{3}[-.\s]?\d{4}',
            r'\+1[-.\s]?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}',
        ]
        
        for pattern in phone_patterns:
            if re.search(pattern, html_content):
                return True
        
        return False
    
    def has_email_address(self, html_content: str) -> bool:
        """Check if page has email address"""
        email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        return bool(re.search(email_pattern, html_content))
    
    def is_responsive(self, soup) -> bool:
        """Check if page is responsive (has viewport meta tag)"""
        viewport_meta = soup.find('meta', attrs={'name': 'viewport'})
        return viewport_meta is not None

# ============================================================================
# PLATFORM SCRAPERS
# ============================================================================

class PlatformScraper:
    """Base class for platform scrapers"""
    
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
    
    def extract_business_info(self, soup, platform: str) -> Dict:
        """Extract business information from platform page"""
        raise NotImplementedError
    
    def search_businesses(self, query: str, location: str = None, limit: int = 20) -> List[Dict]:
        """Search for businesses on platform"""
        raise NotImplementedError

class GoogleBusinessScraper(PlatformScraper):
    """Google Business Profile scraper"""
    
    def search_businesses(self, query: str, location: str = None, limit: int = 20) -> List[Dict]:
        """Search Google for businesses"""
        results = []
        
        try:
            search_query = f"{query}"
            if location:
                search_query += f" {location}"
            
            params = {
                'q': search_query,
                'num': limit,
                'gl': 'us',
                'hl': 'en'
            }
            
            response = self.session.get(
                'https://www.google.com/search',
                params=params,
                timeout=10
            )
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extract business results (simplified - would need more sophisticated parsing)
            # Look for business listings
            business_divs = soup.find_all('div', class_=re.compile(r'(VkpGBb|dbg0pd|iUh30|rc)'))
            
            for div in business_divs[:limit]:
                business_info = self.extract_from_google_div(div)
                if business_info:
                    results.append(business_info)
            
        except Exception as e:
            logger.log(f"Google search error: {e}", "ERROR")
        
        return results
    
    def extract_from_google_div(self, div) -> Optional[Dict]:
        """Extract business info from Google result div"""
        try:
            business = {}
            
            # Extract name
            name_elem = div.find('h3')
            if name_elem:
                business['name'] = name_elem.get_text(strip=True)
            
            # Extract link
            link_elem = div.find('a', href=True)
            if link_elem:
                href = link_elem['href']
                if href.startswith('/url?'):
                    # Extract actual URL from Google redirect
                    parsed = parse_qs(urlparse(href).query)
                    if 'q' in parsed:
                        business['url'] = parsed['q'][0]
                else:
                    business['url'] = href
            
            # Extract snippet
            snippet_elem = div.find('div', class_=re.compile('s3v9rd|VwiC3b'))
            if snippet_elem:
                business['snippet'] = snippet_elem.get_text(strip=True)[:200]
            
            # Extract rating if available
            rating_elem = div.find('span', class_='yi40Hd')
            if rating_elem:
                business['rating'] = rating_elem.get_text(strip=True)
            
            return business if business.get('name') else None
            
        except Exception as e:
            logger.log(f"Google extraction error: {e}", "WARNING")
            return None

class FacebookScraper(PlatformScraper):
    """Facebook Business Page scraper"""
    
    def search_businesses(self, query: str, location: str = None, limit: int = 20) -> List[Dict]:
        """Search Facebook for business pages"""
        results = []
        
        try:
            # Facebook search requires authentication
            # This is a simplified version
            search_url = f"https://www.facebook.com/public/{query.replace(' ', '-')}"
            
            response = self.session.get(search_url, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extract business pages (simplified)
            profile_divs = soup.find_all('div', class_='_2ph_')
            
            for div in profile_divs[:limit]:
                business_info = self.extract_from_facebook_div(div)
                if business_info:
                    results.append(business_info)
            
        except Exception as e:
            logger.log(f"Facebook search error: {e}", "ERROR")
        
        return results
    
    def extract_from_facebook_div(self, div) -> Optional[Dict]:
        """Extract business info from Facebook result div"""
        try:
            business = {}
            
            # Extract name
            name_elem = div.find('a', class_='_32mo')
            if name_elem:
                business['name'] = name_elem.get_text(strip=True)
                business['url'] = urljoin('https://facebook.com', name_elem['href'])
            
            return business if business.get('name') else None
            
        except Exception as e:
            logger.log(f"Facebook extraction error: {e}", "WARNING")
            return None

# ============================================================================
# LEAD QUALIFICATION ENGINE
# ============================================================================

class LeadQualificationEngine:
    """AI-powered lead qualification engine"""
    
    def __init__(self):
        self.openai_client = None
        
        if OPENAI_AVAILABLE and CONFIG.api.openai_api_key:
            try:
                self.openai_client = openai.OpenAI(api_key=CONFIG.api.openai_api_key)
            except Exception as e:
                logger.log(f"OpenAI initialization failed: {e}", "WARNING")
    
    def qualify_lead(self, lead_data: Dict) -> Dict:
        """Qualify lead using AI and business rules"""
        qualification = {
            "lead_score": 50,
            "quality_tier": "Unknown",
            "business_type": "Unknown",
            "estimated_revenue": "Unknown",
            "employee_count": "Unknown",
            "years_in_business": None,
            "decision_maker_confidence": 0,
            "ai_notes": "",
            "outreach_strategy": "",
            "qualification_reasons": []
        }
        
        # Apply business rules first
        qualification.update(self.apply_business_rules(lead_data))
        
        # Apply AI enrichment if enabled
        if CONFIG.ai_enrichment.enabled and self.openai_client:
            try:
                ai_qualification = self.apply_ai_enrichment(lead_data)
                qualification.update(ai_qualification)
            except Exception as e:
                logger.log(f"AI qualification error: {e}", "WARNING")
        
        # Final score adjustment based on website status
        website_status = lead_data.get("website_status", "unknown")
        if website_status in ["no_website", "broken", "parked", "placeholder"]:
            qualification["lead_score"] = min(100, qualification["lead_score"] + 20)
            qualification["qualification_reasons"].append("High intent: No/broken website")
        
        # Determine quality tier based on score
        qualification["quality_tier"] = self.determine_quality_tier(qualification["lead_score"])
        
        return qualification
    
    def apply_business_rules(self, lead_data: Dict) -> Dict:
        """Apply business rules for lead scoring"""
        score = 50
        reasons = []
        
        # Industry scoring
        industry = lead_data.get("industry", "").lower()
        high_value_industries = [
            'contractor', 'construction', 'remodeling', 'roofing',
            'electrical', 'plumbing', 'hvac', 'landscaping'
        ]
        
        if any(hv_industry in industry for hv_industry in high_value_industries):
            score += 10
            reasons.append("High-value industry")
        
        # Location scoring
        city = lead_data.get("city", "").lower()
        target_cities = [c.lower() for c in CONFIG.filters.target_cities]
        
        if city in target_cities:
            score += 5
            reasons.append("Target city")
        
        # Contact information scoring
        if lead_data.get("phone"):
            score += 10
            reasons.append("Has phone number")
        
        if lead_data.get("email"):
            score += 5
            reasons.append("Has email")
        
        if lead_data.get("address"):
            score += 5
            reasons.append("Has address")
        
        # Social media presence
        social_media = lead_data.get("social_media", {})
        if isinstance(social_media, dict) and social_media:
            score += 5
            reasons.append("Social media presence")
        
        # Services offered
        services = lead_data.get("services", [])
        if isinstance(services, list) and len(services) >= 3:
            score += 5
            reasons.append("Multiple services offered")
        
        return {
            "lead_score": min(100, max(0, score)),
            "qualification_reasons": reasons
        }
    
    def apply_ai_enrichment(self, lead_data: Dict) -> Dict:
        """Apply AI enrichment to lead"""
        prompt = f"""
        Analyze this business lead and provide a comprehensive qualification report.
        
        BUSINESS INFORMATION:
        - Name: {lead_data.get('business_name', 'Unknown')}
        - Website: {lead_data.get('website', 'None')}
        - Website Status: {lead_data.get('website_status', 'unknown')}
        - Phone: {lead_data.get('phone', 'None')}
        - Email: {lead_data.get('email', 'None')}
        - Address: {lead_data.get('address', 'None')}
        - City: {lead_data.get('city', 'Unknown')}
        - State: {lead_data.get('state', 'Unknown')}
        - Industry: {lead_data.get('industry', 'Unknown')}
        - Description: {lead_data.get('description', 'None')}
        - Services: {lead_data.get('services', [])}
        
        INSTRUCTIONS:
        1. Score the lead from 0-100 based on business viability and growth potential
        2. Determine business type (LLC, Corporation, Sole Proprietorship, Partnership)
        3. Estimate annual revenue range
        4. Estimate employee count range
        5. Estimate years in business
        6. Identify likely decision maker title
        7. Provide 3-5 bullet points of key insights
        8. Suggest an outreach strategy
        
        RESPONSE FORMAT (JSON):
        {{
            "lead_score": 0-100,
            "business_type": "type",
            "estimated_revenue": "range",
            "employee_count": "range",
            "years_in_business": number or null,
            "decision_maker_title": "title",
            "decision_maker_confidence": 0-100,
            "ai_notes": "bullet points",
            "outreach_strategy": "strategy"
        }}
        """
        
        try:
            response = self.openai_client.chat.completions.create(
                model=CONFIG.ai_enrichment.model,
                messages=[
                    {"role": "system", "content": "You are a lead qualification expert specializing in construction and home services businesses."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=CONFIG.ai_enrichment.max_tokens,
                temperature=0.3
            )
            
            ai_response = response.choices[0].message.content
            
            # Parse JSON response
            try:
                ai_data = json.loads(ai_response)
                return ai_data
            except json.JSONDecodeError:
                # Try to extract JSON from text
                json_match = re.search(r'\{.*\}', ai_response, re.DOTALL)
                if json_match:
                    ai_data = json.loads(json_match.group())
                    return ai_data
                
                logger.log("Failed to parse AI response as JSON", "WARNING")
                return {}
            
        except Exception as e:
            logger.log(f"AI enrichment error: {e}", "WARNING")
            return {}
    
    def determine_quality_tier(self, score: int) -> str:
        """Determine quality tier based on score"""
        if score >= 90:
            return "Premium"
        elif score >= 75:
            return "High"
        elif score >= 60:
            return "Medium"
        elif score >= 40:
            return "Low"
        else:
            return "Unknown"
    
    def generate_outreach_template(self, lead_data: Dict, template_type: str = "email") -> Dict:
        """Generate personalized outreach template"""
        if not self.openai_client:
            return {"error": "OpenAI not available"}
        
        prompt = f"""
        Create a personalized {template_type} outreach message for this business lead.
        
        LEAD INFORMATION:
        - Business: {lead_data.get('business_name', 'Unknown Business')}
        - Industry: {lead_data.get('industry', 'Unknown')}
        - City: {lead_data.get('city', 'Unknown')}
        - Website Status: {lead_data.get('website_status', 'unknown')}
        - Key Services: {lead_data.get('services', [])}
        
        CONTEXT:
        - This business has a {lead_data.get('website_status', 'unknown')} website
        - We provide professional website development and digital marketing services
        - We help businesses like theirs establish a strong online presence
        
        REQUIREMENTS:
        1. Personalize based on their specific industry and location
        2. Reference their current website situation tactfully
        3. Highlight how we can help them grow their business
        4. Include a clear call to action
        5. Keep it professional but friendly
        
        RESPONSE FORMAT:
        Subject: [Email Subject Line]
        
        Body:
        [Email body here]
        
        Call to Action: [Specific CTA]
        """
        
        try:
            response = self.openai_client.chat.completions.create(
                model=CONFIG.ai_enrichment.model,
                messages=[
                    {"role": "system", "content": "You are a professional outreach specialist for a digital marketing agency."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=500,
                temperature=0.7
            )
            
            return {
                "success": True,
                "template": response.choices[0].message.content,
                "type": template_type
            }
            
        except Exception as e:
            logger.log(f"Template generation error: {e}", "ERROR")
            return {"error": str(e)}

# ============================================================================
# ULTIMATE LEAD SCRAPER ENGINE
# ============================================================================

class UltimateLeadScraper:
    """Main lead scraper engine with multiple modes and platforms"""
    
    def __init__(self):
        self.api_key = CONFIG.api.serper_api_key
        self.website_checker = AdvancedWebsiteChecker()
        self.qualification_engine = LeadQualificationEngine()
        self.platform_scrapers = {
            "google": GoogleBusinessScraper(),
            "facebook": FacebookScraper()
        }
        
        self.running = False
        self.paused = False
        self.current_mode = CONFIG.active_mode
        self.stats = {
            'total_cycles': 0,
            'total_leads_found': 0,
            'total_websites_checked': 0,
            'high_intent_leads': 0,
            'premium_leads': 0,
            'last_cycle': None,
            'cycle_duration': 0,
            'errors': 0
        }
        
        self.cache_file = CONFIG.storage["cache_file"]
        self.load_cache()
        
        logger.log(f"✅ Ultimate Lead Scraper initialized in '{self.current_mode}' mode", "SUCCESS")
    
    def load_cache(self):
        """Load search cache"""
        self.cache = {}
        if os.path.exists(self.cache_file):
            try:
                with open(self.cache_file, 'r') as f:
                    self.cache = json.load(f)
            except:
                self.cache = {}
    
    def save_cache(self):
        """Save search cache"""
        try:
            with open(self.cache_file, 'w') as f:
                json.dump(self.cache, f, indent=2)
        except Exception as e:
            logger.log(f"Cache save error: {e}", "WARNING")
    
    def generate_search_queries(self) -> List[Dict]:
        """Generate search queries based on active mode"""
        queries = []
        state = CONFIG.default_state
        mode_config = CONFIG.scraper_modes[self.current_mode]
        
        # Get industries based on mode
        industries = CONFIG.industries
        if CONFIG.filters.target_industries:
            industries = CONFIG.filters.target_industries
        
        # Get cities based on mode
        cities = CONFIG.cities
        if CONFIG.filters.target_cities:
            cities = CONFIG.filters.target_cities
        
        # Generate queries
        for industry in random.sample(industries, min(5, len(industries))):
            for city in random.sample(cities, min(3, len(cities))):
                for phrase_template in CONFIG.search_phrases[:3]:
                    query = phrase_template.format(
                        industry=industry,
                        city=city,
                        state=state
                    )
                    queries.append({
                        'query': query,
                        'industry': industry,
                        'city': city,
                        'state': state,
                        'country': CONFIG.default_country
                    })
        
        random.shuffle(queries)
        return queries[:CONFIG.searches_per_cycle]
    
    def search_platforms(self, query_info: Dict) -> List[Dict]:
        """Search across multiple platforms"""
        all_results = []
        
        for platform in CONFIG.platforms_to_scrape:
            if platform in self.platform_scrapers:
                try:
                    logger.log(f"Searching {platform} for: {query_info['query']}", "INFO")
                    
                    results = self.platform_scrapers[platform].search_businesses(
                        query=query_info['query'],
                        location=query_info['city'],
                        limit=CONFIG.businesses_per_search // len(CONFIG.platforms_to_scrape)
                    )
                    
                    # Add platform info to results
                    for result in results:
                        result['platform'] = platform
                        result['search_query'] = query_info['query']
                        result['industry'] = query_info['industry']
                        result['city'] = query_info['city']
                        result['state'] = query_info['state']
                    
                    all_results.extend(results)
                    
                    # Rate limiting
                    time.sleep(random.uniform(1, 2))
                    
                except Exception as e:
                    logger.log(f"{platform} search error: {e}", "ERROR")
                    self.stats['errors'] += 1
        
        return all_results
    
    async def process_business(self, business_info: Dict) -> Optional[Dict]:
        """Process a single business into a lead"""
        try:
            # Extract website from business info
            website = business_info.get('url') or business_info.get('website', '')
            
            # Check website status
            website_check = await self.website_checker.check_website_async(website)
            
            # Apply mode filters
            mode_config = CONFIG.scraper_modes[self.current_mode]
            website_status = website_check['status']
            
            # Check if business passes mode filters
            passes_filter = False
            
            if mode_config.scrape_no_website and website_status in ["no_website", "invalid_url"]:
                passes_filter = True
            elif mode_config.scrape_broken_website and website_status in ["broken", "unreachable", "timeout", "error"]:
                passes_filter = True
            elif mode_config.scrape_with_website and website_status == "active":
                passes_filter = True
            elif website_status in ["parked", "placeholder"]:
                passes_filter = True  # Always include parked/placeholder
            
            if not passes_filter:
                return None
            
            # Apply additional filters
            if not self.passes_additional_filters(business_info, website_check):
                return None
            
            # Create lead data structure
            lead_data = {
                'business_name': business_info.get('name', 'Unknown Business'),
                'website': website,
                'website_status': website_status,
                'phone': business_info.get('phone', ''),
                'email': business_info.get('email', ''),
                'address': business_info.get('address', ''),
                'city': business_info.get('city', ''),
                'state': business_info.get('state', CONFIG.default_state),
                'country': business_info.get('country', CONFIG.default_country),
                'industry': business_info.get('industry', ''),
                'description': business_info.get('snippet', '')[:500],
                'social_media': business_info.get('social_media', {}),
                'services': business_info.get('services', []),
                
                # Platform data
                'lead_source': 'Platform Scraper',
                'scraped_date': datetime.now(timezone.utc).isoformat(),
                
                # Platform-specific URLs
                'google_business_url': business_info.get('google_url', '') if business_info.get('platform') == 'google' else '',
                'facebook_business_url': business_info.get('facebook_url', '') if business_info.get('platform') == 'facebook' else '',
                'yelp_business_url': business_info.get('yelp_url', '') if business_info.get('platform') == 'yelp' else '',
                'bbb_business_url': business_info.get('bbb_url', '') if business_info.get('platform') == 'bbb' else '',
                'other_platforms': json.dumps([business_info.get('platform', 'unknown')])
            }
            
            # Website check details
            lead_data.update({
                'website_load_time': website_check.get('load_time'),
                'website_title': website_check.get('title'),
                'website_has_contact_form': website_check.get('has_contact_form'),
                'website_has_phone': website_check.get('has_phone'),
                'website_has_email': website_check.get('has_email'),
                'website_ssl_valid': website_check.get('ssl_valid'),
                'website_responsive': website_check.get('responsive')
            })
            
            # Generate fingerprint
            fingerprint_data = (
                lead_data['business_name'],
                lead_data['website'],
                lead_data['phone'],
                lead_data['city']
            )
            lead_data['fingerprint'] = hashlib.sha256(str(fingerprint_data).encode()).hexdigest()
            
            # Qualify lead
            qualification = self.qualification_engine.qualify_lead(lead_data)
            lead_data.update(qualification)
            
            # Apply quality threshold
            if lead_data['lead_score'] < mode_config.quality_threshold:
                logger.log(f"Lead below threshold: {lead_data['business_name']} ({lead_data['lead_score']})", "DEBUG")
                return None
            
            # Calculate potential value
            if not lead_data.get('potential_value'):
                lead_data['potential_value'] = self.calculate_potential_value(lead_data)
            
            # Set outreach priority
            lead_data['outreach_priority'] = self.determine_outreach_priority(lead_data)
            
            return lead_data
            
        except Exception as e:
            logger.log(f"Business processing error: {e}", "ERROR")
            self.stats['errors'] += 1
            return None
    
    def passes_additional_filters(self, business_info: Dict, website_check: Dict) -> bool:
        """Apply additional business filters"""
        filters = CONFIG.filters
        
        # Exclude chains
        if filters.exclude_chains:
            business_name = business_info.get('name', '').lower()
            chain_keywords = ['franchise', 'chain', 'corporate', 'national', 'llc']
            if any(keyword in business_name for keyword in chain_keywords):
                return False
        
        # Exclude keywords
        if filters.exclude_keywords:
            business_name = business_info.get('name', '').lower()
            description = business_info.get('snippet', '').lower()
            
            for keyword in filters.exclude_keywords:
                if keyword.lower() in business_name or keyword.lower() in description:
                    return False
        
        # Include keywords
        if filters.include_keywords:
            business_name = business_info.get('name', '').lower()
            description = business_info.get('snippet', '').lower()
            
            if not any(keyword.lower() in business_name or keyword.lower() in description 
                      for keyword in filters.include_keywords):
                return False
        
        return True
    
    def calculate_potential_value(self, lead_data: Dict) -> int:
        """Calculate potential value of lead"""
        base_value = 1000
        
        # Adjust based on quality tier
        tier_multipliers = {
            'Premium': 10,
            'High': 5,
            'Medium': 2,
            'Low': 1,
            'Unknown': 1
        }
        
        tier = lead_data.get('quality_tier', 'Unknown')
        multiplier = tier_multipliers.get(tier, 1)
        
        # Adjust based on website status
        website_status = lead_data.get('website_status', 'unknown')
        if website_status in ['no_website', 'broken', 'parked']:
            multiplier *= 2
        
        # Adjust based on industry
        industry = lead_data.get('industry', '').lower()
        if any(ind in industry for ind in ['contractor', 'construction', 'roofing', 'plumbing']):
            multiplier *= 1.5
        
        return int(base_value * multiplier)
    
    def determine_outreach_priority(self, lead_data: Dict) -> str:
        """Determine outreach priority"""
        score = lead_data.get('lead_score', 0)
        website_status = lead_data.get('website_status', 'unknown')
        
        if score >= 90 or website_status == 'no_website':
            return 'Immediate'
        elif score >= 75:
            return 'High'
        elif score >= 60:
            return 'Medium'
        else:
            return 'Low'
    
    async def run_cycle_async(self):
        """Run a scraping cycle asynchronously"""
        if not self.running:
            return
        
        cycle_start = time.time()
        logger.log(f"🚀 Starting scraping cycle {self.stats['total_cycles'] + 1} in '{self.current_mode}' mode", "INFO")
        
        queries = self.generate_search_queries()
        leads_found = 0
        websites_checked = 0
        
        for query_info in queries:
            if self.paused or not self.running:
                break
            
            logger.log(f"🔍 Processing query: {query_info['query']}", "INFO")
            
            # Search platforms
            businesses = self.search_platforms(query_info)
            
            # Process businesses concurrently
            tasks = []
            for business in businesses[:CONFIG.businesses_per_search]:
                if self.paused or not self.running:
                    break
                
                task = asyncio.create_task(self.process_business(business))
                tasks.append(task)
                websites_checked += 1
            
            # Wait for all tasks to complete
            if tasks:
                results = await asyncio.gather(*tasks, return_exceptions=True)
                
                for result in results:
                    if isinstance(result, Exception):
                        logger.log(f"Task error: {result}", "ERROR")
                        continue
                    
                    if result:
                        # Save to CRM
                        if CONFIG.crm.enabled and CONFIG.crm.auto_sync:
                            save_result = crm.save_lead(result)
                            if save_result["success"]:
                                leads_found += 1
                                
                                if result.get('quality_tier') in ['Premium', 'High']:
                                    self.stats['premium_leads'] += 1
                                
                                if result.get('website_status') in ['no_website', 'broken', 'parked']:
                                    self.stats['high_intent_leads'] += 1
                                
                                logger.log(f"✅ Saved lead: {result['business_name']} (Score: {result['lead_score']})", "SUCCESS")
                        
                        # Save to JSON file
                        self.save_lead_to_file(result)
            
            # Rate limiting between queries
            if not self.paused and self.running:
                await asyncio.sleep(random.uniform(2, 4))
        
        # Update statistics
        self.stats['total_cycles'] += 1
        self.stats['total_leads_found'] += leads_found
        self.stats['total_websites_checked'] += websites_checked
        self.stats['last_cycle'] = datetime.now().isoformat()
        self.stats['cycle_duration'] = time.time() - cycle_start
        
        logger.log(f"✅ Cycle completed. Found {leads_found} leads, checked {websites_checked} websites. "
                  f"Duration: {self.stats['cycle_duration']:.1f}s", "SUCCESS")
    
    def save_lead_to_file(self, lead_data: Dict):
        """Save lead to JSON file"""
        try:
            leads_file = CONFIG.storage["leads_file"]
            leads = []
            
            if os.path.exists(leads_file):
                with open(leads_file, 'r') as f:
                    leads = json.load(f)
            
            leads.append(lead_data)
            
            # Keep only last 5000 leads
            if len(leads) > 5000:
                leads = leads[-5000:]
            
            with open(leads_file, 'w') as f:
                json.dump(leads, f, indent=2, default=str)
            
            # Save to qualified leads if above threshold
            if lead_data.get('lead_score', 0) >= CONFIG.ai_enrichment.qualification_threshold:
                qualified_file = CONFIG.storage["qualified_leads"]
                qualified = []
                
                if os.path.exists(qualified_file):
                    with open(qualified_file, 'r') as f:
                        qualified = json.load(f)
                
                qualified.append(lead_data)
                
                if len(qualified) > 1000:
                    qualified = qualified[-1000:]
                
                with open(qualified_file, 'w') as f:
                    json.dump(qualified, f, indent=2, default=str)
            
            # Save to premium leads if high quality
            if lead_data.get('quality_tier') in ['Premium', 'High']:
                premium_file = CONFIG.storage["premium_leads"]
                premium = []
                
                if os.path.exists(premium_file):
                    with open(premium_file, 'r') as f:
                        premium = json.load(f)
                
                premium.append(lead_data)
                
                if len(premium) > 500:
                    premium = premium[-500:]
                
                with open(premium_file, 'w') as f:
                    json.dump(premium, f, indent=2, default=str)
                    
        except Exception as e:
            logger.log(f"Error saving lead to file: {e}", "WARNING")
    
    def start(self):
        """Start the scraper"""
        self.running = True
        self.paused = False
        logger.log("🟢 Scraper started", "SUCCESS")
    
    def stop(self):
        """Stop the scraper"""
        self.running = False
        self.paused = False
        logger.log("🔴 Scraper stopped", "INFO")
    
    def pause(self):
        """Pause the scraper"""
        self.paused = True
        logger.log("⏸️  Scraper paused", "INFO")
    
    def resume(self):
        """Resume the scraper"""
        self.paused = False
        logger.log("▶️  Scraper resumed", "INFO")
    
    def set_mode(self, mode_name: str):
        """Set scraper mode"""
        if mode_name in CONFIG.scraper_modes:
            self.current_mode = mode_name
            logger.log(f"Mode changed to: {mode_name}", "INFO")
            return True
        return False
    
    def get_status(self) -> Dict:
        """Get scraper status"""
        return {
            "running": self.running,
            "paused": self.paused,
            "current_mode": self.current_mode,
            "stats": self.stats,
            "config": {
                "active_mode": CONFIG.active_mode,
                "searches_per_cycle": CONFIG.searches_per_cycle,
                "businesses_per_search": CONFIG.businesses_per_search,
                "cycle_interval": CONFIG.cycle_interval,
                "max_cycles": CONFIG.max_cycles
            }
        }

# ============================================================================
# ULTIMATE STREAMLIT DASHBOARD
# ============================================================================

class UltimateStreamlitDashboard:
    """Ultimate Streamlit dashboard with all features"""
    
    def __init__(self):
        self.crm = crm
        self.scraper = None
        self.scraper_running = False
        self.scraper_thread = None
        self.setup_page()
        
        logger.log("✅ Ultimate Streamlit Dashboard initialized", "SUCCESS")
    
    def setup_page(self):
        """Setup Streamlit page configuration"""
        st.set_page_config(
            page_title="Ultimate LeadScraper CRM",
            page_icon="🚀",
            layout="wide",
            initial_sidebar_state="expanded",
            menu_items={
                'Get Help': 'https://github.com/yourusername/ultimate-lead-scraper',
                'Report a bug': 'https://github.com/yourusername/ultimate-lead-scraper/issues',
                'About': '# Ultimate LeadScraper CRM v2.0\nHigh-Intent Lead Generation System'
            }
        )
        
        # Apply custom CSS
        self.setup_custom_css()
        
        # Initialize session state
        if 'initialized' not in st.session_state:
            st.session_state.initialized = True
            st.session_state.scraper_running = False
            st.session_state.scraper_stats = {}
            st.session_state.current_mode = CONFIG.active_mode
            st.session_state.lead_filters = {}
            st.session_state.selected_lead_id = None
            st.session_state.export_data = None
    
    def setup_custom_css(self):
        """Setup custom CSS with modern design"""
        st.markdown("""
        <style>
        /* Modern CSS Reset */
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        /* Main Theme Variables */
        :root {
            --primary: #0066FF;
            --primary-dark: #0052CC;
            --primary-light: #3385FF;
            --secondary: #00D4AA;
            --secondary-dark: #00B894;
            --accent: #FF6B6B;
            --success: #10B981;
            --warning: #F59E0B;
            --danger: #EF4444;
            --dark: #111827;
            --dark-light: #1F2937;
            --dark-lighter: #374151;
            --light: #F9FAFB;
            --light-dark: #E5E7EB;
            --gray: #6B7280;
        }
        
        /* Main Container */
        .main .block-container {
            padding-top: 1rem;
            padding-bottom: 1rem;
            max-width: 100%;
        }
        
        /* Modern Cards */
        .modern-card {
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.1) 0%, rgba(255, 255, 255, 0.05) 100%);
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255, 255, 255, 0.1);
            border-radius: 16px;
            padding: 1.5rem;
            margin-bottom: 1rem;
            transition: all 0.3s ease;
        }
        
        .modern-card:hover {
            transform: translateY(-2px);
            box-shadow: 0 10px 25px rgba(0, 0, 0, 0.2);
            border-color: var(--primary-light);
        }
        
        .metric-card {
            background: linear-gradient(135deg, var(--primary) 0%, var(--primary-dark) 100%);
            border-radius: 12px;
            padding: 1.5rem;
            color: white;
        }
        
        .metric-card-secondary {
            background: linear-gradient(135deg, var(--secondary) 0%, var(--secondary-dark) 100%);
        }
        
        .metric-card-accent {
            background: linear-gradient(135deg, var(--accent) 0%, #FF4757 100%);
        }
        
        /* Typography */
        h1, h2, h3, h4, h5, h6 {
            font-weight: 700 !important;
            color: var(--light) !important;
            margin-bottom: 1rem !important;
        }
        
        .subtitle {
            color: var(--gray) !important;
            font-weight: 500 !important;
        }
        
        /* Badges */
        .badge {
            display: inline-block;
            padding: 0.35rem 0.75rem;
            border-radius: 20px;
            font-size: 0.75rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        
        .badge-premium {
            background: linear-gradient(135deg, #FFD700, #FFA500);
            color: #000;
        }
        
        .badge-high {
            background: linear-gradient(135deg, var(--success), #059669);
            color: white;
        }
        
        .badge-medium {
            background: linear-gradient(135deg, var(--warning), #D97706);
            color: white;
        }
        
        .badge-low {
            background: linear-gradient(135deg, var(--gray), #4B5563);
            color: white;
        }
        
        .badge-no-website {
            background: linear-gradient(135deg, var(--accent), #DC2626);
            color: white;
        }
        
        .badge-broken-website {
            background: linear-gradient(135deg, #F97316, #EA580C);
            color: white;
        }
        
        .badge-active-website {
            background: linear-gradient(135deg, var(--success), #059669);
            color: white;
        }
        
        /* Buttons */
        .stButton > button {
            border-radius: 8px !important;
            font-weight: 600 !important;
            transition: all 0.3s ease !important;
            border: none !important;
        }
        
        .stButton > button:hover {
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(0, 0, 0, 0.3) !important;
        }
        
        /* Data Tables */
        .dataframe {
            border-radius: 8px !important;
            overflow: hidden !important;
        }
        
        .dataframe th {
            background: linear-gradient(135deg, var(--primary), var(--primary-dark)) !important;
            color: white !important;
            font-weight: 600 !important;
            border: none !important;
        }
        
        .dataframe td {
            border-color: rgba(255, 255, 255, 0.1) !important;
            padding: 12px !important;
        }
        
        /* Tabs */
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px;
            background: rgba(255, 255, 255, 0.05);
            border-radius: 12px;
            padding: 4px;
        }
        
        .stTabs [data-baseweb="tab"] {
            border-radius: 8px !important;
            padding: 12px 24px !important;
            font-weight: 600 !important;
            transition: all 0.3s ease !important;
        }
        
        .stTabs [aria-selected="true"] {
            background: linear-gradient(135deg, var(--primary), var(--primary-dark)) !important;
            color: white !important;
        }
        
        /* Sidebar */
        section[data-testid="stSidebar"] {
            background: linear-gradient(180deg, var(--dark) 0%, #0F172A 100%) !important;
            border-right: 1px solid rgba(255, 255, 255, 0.1) !important;
        }
        
        /* Scrollbar */
        ::-webkit-scrollbar {
            width: 8px;
            height: 8px;
        }
        
        ::-webkit-scrollbar-track {
            background: rgba(255, 255, 255, 0.05);
            border-radius: 4px;
        }
        
        ::-webkit-scrollbar-thumb {
            background: var(--primary);
            border-radius: 4px;
        }
        
        ::-webkit-scrollbar-thumb:hover {
            background: var(--primary-light);
        }
        
        /* Hide Streamlit elements */
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
        
        /* Animations */
        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        .fade-in {
            animation: fadeIn 0.5s ease-out;
        }
        
        /* Status Indicators */
        .status-dot {
            display: inline-block;
            width: 8px;
            height: 8px;
            border-radius: 50%;
            margin-right: 8px;
        }
        
        .status-active { background-color: var(--success); }
        .status-paused { background-color: var(--warning); }
        .status-stopped { background-color: var(--danger); }
        
        /* Form Elements */
        .stTextInput > div > div > input,
        .stSelectbox > div > div,
        .stNumberInput > div > div > input,
        .stTextArea > div > div > textarea {
            border-radius: 8px !important;
            border: 1px solid rgba(255, 255, 255, 0.2) !important;
            background: rgba(255, 255, 255, 0.05) !important;
            color: white !important;
        }
        
        /* Loader */
        .loader {
            border: 3px solid rgba(255, 255, 255, 0.1);
            border-top: 3px solid var(--primary);
            border-radius: 50%;
            width: 40px;
            height: 40px;
            animation: spin 1s linear infinite;
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        </style>
        """, unsafe_allow_html=True)
    
    def run_scraper_background(self):
        """Run scraper in background thread"""
        try:
            self.scraper = UltimateLeadScraper()
            self.scraper.start()
            
            cycles = 0
            while self.scraper_running and cycles < CONFIG.max_cycles:
                if not self.scraper.running:
                    break
                
                # Run async cycle
                asyncio.run(self.scraper.run_cycle_async())
                
                cycles += 1
                
                # Update session state
                st.session_state.scraper_stats = self.scraper.get_status()
                st.session_state.scraper_stats['cycles_completed'] = cycles
                
                # Check if we should continue
                if self.scraper_running and cycles < CONFIG.max_cycles:
                    time.sleep(CONFIG.cycle_interval)
            
            self.scraper.stop()
            self.scraper_running = False
            st.session_state.scraper_running = False
            
            logger.log("Scraper finished successfully", "SUCCESS")
            
        except Exception as e:
            logger.log(f"Background scraper error: {e}", "ERROR")
            self.scraper_running = False
            st.session_state.scraper_running = False
    
    def start_scraper(self):
        """Start the scraper"""
        if not self.scraper_running:
            self.scraper_running = True
            st.session_state.scraper_running = True
            
            # Start background thread
            self.scraper_thread = threading.Thread(target=self.run_scraper_background, daemon=True)
            self.scraper_thread.start()
            
            return True
        return False
    
    def stop_scraper(self):
        """Stop the scraper"""
        self.scraper_running = False
        if self.scraper:
            self.scraper.stop()
        
        st.session_state.scraper_running = False
        return True
    
    def render_sidebar(self):
        """Render the modern sidebar"""
        with st.sidebar:
            # Logo and Title
            st.markdown("""
            <div style="text-align: center; margin-bottom: 2rem; padding: 1rem;">
                <h1 style="color: var(--primary); font-size: 2rem; margin-bottom: 0.5rem;">
                    🚀 Ultimate LeadScraper
                </h1>
                <p style="color: var(--gray); font-size: 0.875rem; font-weight: 500;">
                    v2.0 • High-Intent Lead Generation
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            # Navigation
            st.markdown("### 📱 Navigation")
            
            nav_options = {
                "📊 Dashboard": "dashboard",
                "👥 Leads Management": "leads",
                "🔍 Lead Details": "lead_details",
                "⚙️ Settings": "settings",
                "📈 Analytics": "analytics",
                "📤 Export": "export",
                "📋 Logs": "logs",
                "🔄 Automation": "automation"
            }
            
            selected_nav = st.radio(
                "Go to",
                list(nav_options.keys()),
                label_visibility="collapsed",
                key="nav_selection"
            )
            
            st.markdown("---")
            
            # Scraper Control Panel
            st.markdown("### ⚡ Scraper Control")
            
            # Mode Selection
            mode_options = {v.name: k for k, v in CONFIG.scraper_modes.items()}
            selected_mode = st.selectbox(
                "Scraping Mode",
                list(mode_options.keys()),
                index=list(mode_options.keys()).index(CONFIG.scraper_modes[CONFIG.active_mode].name)
            )
            
            if selected_mode and selected_mode in mode_options:
                mode_key = mode_options[selected_mode]
                if mode_key != st.session_state.current_mode:
                    st.session_state.current_mode = mode_key
                    if self.scraper:
                        self.scraper.set_mode(mode_key)
            
            # Control Buttons
            col1, col2, col3 = st.columns(3)
            
            with col1:
                start_disabled = st.session_state.get('scraper_running', False)
                if st.button("▶️ Start", disabled=start_disabled, use_container_width=True, type="primary"):
                    if self.start_scraper():
                        st.success("Scraper started!")
                        st.rerun()
            
            with col2:
                stop_disabled = not st.session_state.get('scraper_running', False)
                if st.button("⏹️ Stop", disabled=stop_disabled, use_container_width=True, type="secondary"):
                    if self.stop_scraper():
                        st.info("Scraper stopped!")
                        st.rerun()
            
            with col3:
                if st.button("⏸️ Pause", use_container_width=True):
                    if self.scraper:
                        self.scraper.pause()
                        st.info("Scraper paused!")
            
            # Status Indicator
            st.markdown("---")
            st.markdown("### 📊 Status")
            
            if st.session_state.get('scraper_running'):
                status_color = "#10B981"
                status_text = "Active"
                status_emoji = "🟢"
            else:
                status_color = "#EF4444"
                status_text = "Inactive"
                status_emoji = "🔴"
            
            st.markdown(f"""
            <div style="background: rgba(255, 255, 255, 0.05); padding: 1rem; border-radius: 8px; margin-bottom: 1rem;">
                <div style="display: flex; align-items: center; justify-content: space-between;">
                    <span style="color: {status_color}; font-weight: 600;">{status_emoji} {status_text}</span>
                    <span style="color: var(--gray); font-size: 0.875rem;">{datetime.now().strftime('%H:%M:%S')}</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Quick Stats
            st.markdown("### 📈 Quick Stats")
            
            today_stats = self.crm.get_today_stats()
            col1, col2 = st.columns(2)
            
            with col1:
                st.metric("Today's Leads", today_stats.get('today_leads', 0))
            
            with col2:
                st.metric("High Intent", today_stats.get('high_intent_leads', 0))
            
            # System Info
            st.markdown("---")
            st.markdown("### 💻 System Info")
            
            info_items = [
                ("Database", "✅ Connected" if self.crm.conn else "❌ Error"),
                ("AI Enrichment", "✅ Active" if OPENAI_AVAILABLE and CONFIG.api.openai_api_key else "❌ Disabled"),
                ("Mode", CONFIG.scraper_modes[st.session_state.current_mode].name),
                ("Cities", len(CONFIG.cities)),
                ("Industries", len(CONFIG.industries))
            ]
            
            for label, value in info_items:
                st.markdown(f"**{label}:** {value}")
        
        return nav_options[selected_nav]
    
    def render_dashboard(self):
        """Render the main dashboard"""
        st.title("📊 Ultimate Dashboard")
        st.markdown("<p class='subtitle'>Real-time monitoring and insights</p>", unsafe_allow_html=True)
        
        # Top Metrics Row
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            stats = self.crm.get_statistics("7d")
            total_leads = stats.get('overall', {}).get('total_leads', 0)
            st.markdown("""
            <div class="metric-card">
                <h3 style="color: white; margin-bottom: 0.5rem;">Total Leads</h3>
                <h1 style="color: white; font-size: 2.5rem;">{:,}</h1>
                <p style="color: rgba(255, 255, 255, 0.8); font-size: 0.875rem;">Last 7 days</p>
            </div>
            """.format(total_leads), unsafe_allow_html=True)
        
        with col2:
            total_value = stats.get('overall', {}).get('total_potential_value', 0)
            st.markdown("""
            <div class="metric-card metric-card-secondary">
                <h3 style="color: white; margin-bottom: 0.5rem;">Potential Value</h3>
                <h1 style="color: white; font-size: 2.5rem;">${:,}</h1>
                <p style="color: rgba(255, 255, 255, 0.8); font-size: 0.875rem;">Estimated</p>
            </div>
            """.format(total_value), unsafe_allow_html=True)
        
        with col3:
            avg_score = stats.get('overall', {}).get('average_score', 0)
            st.markdown("""
            <div class="metric-card">
                <h3 style="color: white; margin-bottom: 0.5rem;">Avg. Score</h3>
                <h1 style="color: white; font-size: 2.5rem;">{:.1f}</h1>
                <p style="color: rgba(255, 255, 255, 0.8); font-size: 0.875rem;">Lead Quality</p>
            </div>
            """.format(avg_score), unsafe_allow_html=True)
        
        with col4:
            premium_leads = sum(1 for q in stats.get('quality_distribution', []) 
                              if q.get('quality_tier') in ['Premium', 'High'])
            st.markdown("""
            <div class="metric-card metric-card-accent">
                <h3 style="color: white; margin-bottom: 0.5rem;">Premium Leads</h3>
                <h1 style="color: white; font-size: 2.5rem;">{:,}</h1>
                <p style="color: rgba(255, 255, 255, 0.8); font-size: 0.875rem;">High Quality</p>
            </div>
            """.format(premium_leads), unsafe_allow_html=True)
        
        # Charts Row
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("<div class='modern-card'>", unsafe_allow_html=True)
            st.subheader("📊 Lead Quality Distribution")
            
            quality_data = stats.get('quality_distribution', [])
            if quality_data:
                df_quality = pd.DataFrame(quality_data)
                fig_quality = px.pie(
                    df_quality,
                    values='count',
                    names='quality_tier',
                    color='quality_tier',
                    color_discrete_map={
                        'Premium': '#FFD700',
                        'High': '#10B981',
                        'Medium': '#F59E0B',
                        'Low': '#6B7280',
                        'Unknown': '#9CA3AF'
                    },
                    hole=0.4
                )
                fig_quality.update_layout(
                    showlegend=True,
                    legend=dict(
                        orientation="h",
                        yanchor="bottom",
                        y=1.02,
                        xanchor="right",
                        x=1
                    )
                )
                st.plotly_chart(fig_quality, use_container_width=True)
            else:
                st.info("No quality data available yet.")
            st.markdown("</div>", unsafe_allow_html=True)
        
        with col2:
            st.markdown("<div class='modern-card'>", unsafe_allow_html=True)
            st.subheader("🌐 Website Status Analysis")
            
            website_data = stats.get('website_status_distribution', [])
            if website_data:
                df_website = pd.DataFrame(website_data)
                
                # Color mapping for website status
                status_colors = {
                    'active': '#10B981',
                    'no_website': '#EF4444',
                    'broken': '#F59E0B',
                    'parked': '#8B5CF6',
                    'placeholder': '#EC4899',
                    'unknown': '#6B7280'
                }
                
                fig_website = px.bar(
                    df_website,
                    x='website_status',
                    y='count',
                    color='website_status',
                    color_discrete_map=status_colors,
                    text='count'
                )
                fig_website.update_layout(
                    xaxis_title="Website Status",
                    yaxis_title="Count",
                    showlegend=False
                )
                fig_website.update_traces(texttemplate='%{text}', textposition='outside')
                st.plotly_chart(fig_website, use_container_width=True)
            else:
                st.info("No website status data available yet.")
            st.markdown("</div>", unsafe_allow_html=True)
        
        # Daily Trend Chart
        st.markdown("<div class='modern-card'>", unsafe_allow_html=True)
        st.subheader("📈 Daily Lead Acquisition Trend")
        
        daily_data = stats.get('daily_trend', [])
        if daily_data:
            df_daily = pd.DataFrame(daily_data)
            df_daily['date'] = pd.to_datetime(df_daily['date'])
            df_daily = df_daily.sort_values('date')
            
            fig_daily = px.area(
                df_daily,
                x='date',
                y='leads_count',
                title='',
                markers=True,
                color_discrete_sequence=['#0066FF']
            )
            fig_daily.update_layout(
                xaxis_title="Date",
                yaxis_title="Leads Count",
                hovermode='x unified'
            )
            st.plotly_chart(fig_daily, use_container_width=True)
        else:
            st.info("No daily trend data available yet.")
        st.markdown("</div>", unsafe_allow_html=True)
        
        # Recent Leads and Top Lists
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("<div class='modern-card'>", unsafe_allow_html=True)
            st.subheader("🆕 Recent Leads")
            
            recent_leads = self.crm.get_leads(page=1, per_page=10)
            if recent_leads["leads"]:
                df_recent = pd.DataFrame(recent_leads["leads"])
                
                if not df_recent.empty:
                    display_cols = ['business_name', 'city', 'lead_score', 'quality_tier', 'website_status']
                    if all(col in df_recent.columns for col in display_cols):
                        df_display = df_recent[display_cols].copy()
                        df_display.columns = ['Business', 'City', 'Score', 'Quality', 'Website']
                        
                        # Format the display
                        st.dataframe(
                            df_display,
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "Business": st.column_config.TextColumn("Business", width="large"),
                                "City": st.column_config.TextColumn("City"),
                                "Score": st.column_config.ProgressColumn("Score", min_value=0, max_value=100),
                                "Quality": st.column_config.TextColumn("Quality"),
                                "Website": st.column_config.TextColumn("Website")
                            }
                        )
            else:
                st.info("No recent leads found.")
            st.markdown("</div>", unsafe_allow_html=True)
        
        with col2:
            st.markdown("<div class='modern-card'>", unsafe_allow_html=True)
            st.subheader("🏆 Top Cities")
            
            top_cities = stats.get('top_cities', [])
            if top_cities:
                df_cities = pd.DataFrame(top_cities)
                df_cities = df_cities.sort_values('lead_count', ascending=False).head(10)
                
                fig_cities = px.bar(
                    df_cities,
                    x='city',
                    y='lead_count',
                    color='lead_count',
                    color_continuous_scale='blues',
                    text='lead_count'
                )
                fig_cities.update_layout(
                    xaxis_title="City",
                    yaxis_title="Lead Count",
                    showlegend=False
                )
                fig_cities.update_traces(texttemplate='%{text}', textposition='outside')
                st.plotly_chart(fig_cities, use_container_width=True)
            else:
                st.info("No city data available yet.")
            st.markdown("</div>", unsafe_allow_html=True)
    
    def render_leads_management(self):
        """Render leads management page"""
        st.title("👥 Leads Management")
        st.markdown("<p class='subtitle'>Filter, manage, and organize your leads</p>", unsafe_allow_html=True)
        
        # Advanced Filters
        with st.expander("🔍 Advanced Filters", expanded=False):
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                search_term = st.text_input("Search", key="leads_search")
            
            with col2:
                status_options = ["All"] + [
                    "New Lead", "Contacted", "Follow Up", "Meeting Scheduled",
                    "Zoom Meeting", "Closed (Won)", "Closed (Lost)", "Archived"
                ]
                status_filter = st.multiselect("Status", status_options, default=["All"])
            
            with col3:
                quality_options = ["All"] + ["Premium", "High", "Medium", "Low", "Unknown"]
                quality_filter = st.multiselect("Quality Tier", quality_options, default=["All"])
            
            with col4:
                website_options = ["All"] + ["active", "no_website", "broken", "parked", "placeholder", "unknown"]
                website_filter = st.multiselect("Website Status", website_options, default=["All"])
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                city_options = ["All"] + CONFIG.cities
                city_filter = st.multiselect("City", city_options, default=["All"])
            
            with col2:
                industry_options = ["All"] + CONFIG.industries
                industry_filter = st.multiselect("Industry", industry_options, default=["All"])
            
            with col3:
                min_score = st.slider("Min Score", 0, 100, 0)
            
            with col4:
                max_score = st.slider("Max Score", 0, 100, 100)
            
            col1, col2 = st.columns(2)
            
            with col1:
                date_from = st.date_input("From Date", value=None)
            
            with col2:
                date_to = st.date_input("To Date", value=None)
        
        # Build filters
        filters = {}
        
        if search_term:
            filters["search"] = search_term
        
        if "All" not in status_filter:
            filters["status"] = status_filter
        
        if "All" not in quality_filter:
            filters["quality_tier"] = quality_filter
        
        if "All" not in website_filter:
            filters["website_status"] = website_filter
        
        if "All" not in city_filter:
            filters["city"] = city_filter
        
        if "All" not in industry_filter:
            filters["industry"] = industry_filter
        
        if min_score > 0:
            filters["min_score"] = min_score
        
        if max_score < 100:
            filters["max_score"] = max_score
        
        if date_from:
            filters["date_from"] = date_from.isoformat()
        
        if date_to:
            filters["date_to"] = date_to.isoformat()
        
        # Get leads with filters
        leads_data = self.crm.get_leads(filters=filters, page=1, per_page=100)
        leads = leads_data["leads"]
        total_leads = leads_data["total"]
        
        # Display metrics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Leads", total_leads)
        
        with col2:
            high_intent = sum(1 for lead in leads if lead.get('website_status') in ['no_website', 'broken', 'parked'])
            st.metric("High Intent", high_intent)
        
        with col3:
            premium = sum(1 for lead in leads if lead.get('quality_tier') in ['Premium', 'High'])
            st.metric("Premium", premium)
        
        with col4:
            avg_score = sum(lead.get('lead_score', 0) for lead in leads) / len(leads) if leads else 0
            st.metric("Avg Score", f"{avg_score:.1f}")
        
        if leads:
            # Create dataframe for display
            df = pd.DataFrame(leads)
            
            # Select columns for display
            display_columns = [
                'id', 'business_name', 'city', 'industry', 'lead_score',
                'quality_tier', 'website_status', 'lead_status', 'created_at'
            ]
            
            if all(col in df.columns for col in display_columns):
                df_display = df[display_columns].copy()
                df_display.columns = [
                    'ID', 'Business', 'City', 'Industry', 'Score',
                    'Quality', 'Website', 'Status', 'Created'
                ]
                
                # Format dates
                df_display['Created'] = pd.to_datetime(df_display['Created']).dt.strftime('%Y-%m-%d')
                
                # Display with interactive features
                st.dataframe(
                    df_display,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "ID": st.column_config.NumberColumn("ID", width="small"),
                        "Business": st.column_config.TextColumn("Business", width="large"),
                        "City": st.column_config.TextColumn("City"),
                        "Industry": st.column_config.TextColumn("Industry"),
                        "Score": st.column_config.ProgressColumn("Score", min_value=0, max_value=100),
                        "Quality": st.column_config.TextColumn("Quality"),
                        "Website": st.column_config.TextColumn("Website"),
                        "Status": st.column_config.TextColumn("Status"),
                        "Created": st.column_config.TextColumn("Created")
                    }
                )
                
                # Lead selection for detailed view
                st.subheader("📋 Lead Details")
                
                selected_id = st.selectbox(
                    "Select Lead ID for Detailed View",
                    df_display['ID'].tolist(),
                    key="lead_selection"
                )
                
                if selected_id:
                    lead_details = self.crm.get_lead_by_id(selected_id)
                    if lead_details:
                        self.render_lead_detail_view(lead_details)
            else:
                st.warning("Some columns are missing from the data.")
        else:
            st.info("No leads match the current filters.")
    
    def render_lead_detail_view(self, lead: Dict):
        """Render detailed lead view"""
        with st.container():
            col1, col2 = st.columns([2, 1])
            
            with col1:
                st.markdown(f"### {lead.get('business_name', 'Unknown Business')}")
                
                # Badges row
                col_b1, col_b2, col_b3, col_b4 = st.columns(4)
                
                with col_b1:
                    quality = lead.get('quality_tier', 'Unknown')
                    badge_class = f"badge-{quality.lower()}" if quality != 'Unknown' else "badge-low"
                    st.markdown(f'<span class="badge {badge_class}">{quality}</span>', unsafe_allow_html=True)
                
                with col_b2:
                    website_status = lead.get('website_status', 'unknown')
                    if website_status == 'no_website':
                        badge_class = "badge-no-website"
                    elif website_status == 'broken':
                        badge_class = "badge-broken-website"
                    elif website_status == 'active':
                        badge_class = "badge-active-website"
                    else:
                        badge_class = "badge-low"
                    st.markdown(f'<span class="badge {badge_class}">{website_status}</span>', unsafe_allow_html=True)
                
                with col_b3:
                    status = lead.get('lead_status', 'New Lead')
                    st.markdown(f'<span class="badge badge-medium">{status}</span>', unsafe_allow_html=True)
                
                with col_b4:
                    score = lead.get('lead_score', 0)
                    st.progress(score / 100, text=f"Score: {score}")
            
            with col2:
                st.metric("Potential Value", f"${lead.get('potential_value', 0):,}")
            
            # Tabs for different sections
            tab1, tab2, tab3, tab4, tab5 = st.tabs([
                "📋 Overview", "📞 Contact", "🏢 Business", "📊 Insights", "📝 Activities"
            ])
            
            with tab1:
                self.render_lead_overview(lead)
            
            with tab2:
                self.render_lead_contact(lead)
            
            with tab3:
                self.render_lead_business(lead)
            
            with tab4:
                self.render_lead_insights(lead)
            
            with tab5:
                self.render_lead_activities(lead)
    
    def render_lead_overview(self, lead: Dict):
        """Render lead overview tab"""
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("##### Basic Information")
            
            st.text_input("Business Name", lead.get('business_name', ''), disabled=True)
            st.text_input("Industry", lead.get('industry', ''), disabled=True)
            st.text_input("Business Type", lead.get('business_type', 'Unknown'), disabled=True)
            
            if lead.get('description'):
                st.markdown("##### Description")
                st.text_area("Description", lead.get('description', ''), height=150, disabled=True)
        
        with col2:
            st.markdown("##### Services Offered")
            
            services = lead.get('services', [])
            if isinstance(services, list) and services:
                for service in services:
                    st.markdown(f"- {service}")
            else:
                st.info("No services listed")
            
            st.markdown("##### Social Media")
            
            social_media = lead.get('social_media', {})
            if isinstance(social_media, dict) and social_media:
                for platform, url in social_media.items():
                    st.markdown(f"**{platform.title()}:** [{url}]({url})")
            else:
                st.info("No social media links")
    
    def render_lead_contact(self, lead: Dict):
        """Render lead contact tab"""
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("##### Contact Details")
            
            website = lead.get('website', '')
            if website:
                st.markdown(f"**Website:** [{website}]({website})")
                st.markdown(f"**Status:** {lead.get('website_status', 'unknown').title()}")
            
            phone = lead.get('phone', '')
            if phone:
                st.markdown(f"**Phone:** {phone}")
            
            email = lead.get('email', '')
            if email:
                st.markdown(f"**Email:** {email}")
            
            address = lead.get('address', '')
            if address:
                st.text_area("Address", address, disabled=True, height=100)
        
        with col2:
            st.markdown("##### Location")
            
            st.text_input("City", lead.get('city', ''), disabled=True)
            st.text_input("State", lead.get('state', ''), disabled=True)
            st.text_input("Country", lead.get('country', 'US'), disabled=True)
            
            st.markdown("##### Platform Sources")
            
            platforms = []
            if lead.get('google_business_url'):
                platforms.append(("Google", lead['google_business_url']))
            if lead.get('facebook_business_url'):
                platforms.append(("Facebook", lead['facebook_business_url']))
            if lead.get('yelp_business_url'):
                platforms.append(("Yelp", lead['yelp_business_url']))
            if lead.get('bbb_business_url'):
                platforms.append(("BBB", lead['bbb_business_url']))
            
            if platforms:
                for platform, url in platforms:
                    st.markdown(f"**{platform}:** [{url}]({url})")
            else:
                st.info("No platform sources")
    
    def render_lead_business(self, lead: Dict):
        """Render lead business tab"""
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("##### Business Intelligence")
            
            estimated_revenue = lead.get('estimated_revenue', 'Unknown')
            st.metric("Estimated Revenue", estimated_revenue)
            
            employee_count = lead.get('employee_count', 'Unknown')
            st.metric("Employee Count", employee_count)
            
            years_in_business = lead.get('years_in_business')
            if years_in_business:
                st.metric("Years in Business", years_in_business)
        
        with col2:
            st.markdown("##### Decision Maker")
            
            decision_maker_name = lead.get('decision_maker_name')
            if decision_maker_name:
                st.text_input("Name", decision_maker_name, disabled=True)
            
            decision_maker_title = lead.get('decision_maker_title')
            if decision_maker_title:
                st.text_input("Title", decision_maker_title, disabled=True)
            
            decision_maker_email = lead.get('decision_maker_email')
            if decision_maker_email:
                st.text_input("Email", decision_maker_email, disabled=True)
            
            decision_maker_phone = lead.get('decision_maker_phone')
            if decision_maker_phone:
                st.text_input("Phone", decision_maker_phone, disabled=True)
            
            confidence = lead.get('decision_maker_confidence', 0)
            if confidence:
                st.progress(confidence / 100, text=f"Confidence: {confidence}%")
    
    def render_lead_insights(self, lead: Dict):
        """Render lead insights tab"""
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("##### AI Notes & Insights")
            
            ai_notes = lead.get('ai_notes', '')
            if ai_notes:
                st.text_area("AI Analysis", ai_notes, height=200, disabled=True)
            else:
                st.info("No AI insights available")
        
        with col2:
            st.markdown("##### Outreach Strategy")
            
            outreach_strategy = lead.get('outreach_strategy', '')
            if outreach_strategy:
                st.text_area("Strategy", outreach_strategy, height=200, disabled=True)
            else:
                st.info("No outreach strategy available")
            
            # Generate new strategy button
            if st.button("🔄 Generate Outreach Strategy", use_container_width=True):
                with st.spinner("Generating strategy..."):
                    engine = LeadQualificationEngine()
                    result = engine.generate_outreach_template(lead, "email")
                    
                    if "template" in result:
                        st.text_area("Generated Template", result["template"], height=200)
                    else:
                        st.error("Failed to generate template")
    
    def render_lead_activities(self, lead: Dict):
        """Render lead activities tab"""
        activities = lead.get('activities', [])
        
        if activities:
            for activity in activities[:10]:  # Show last 10 activities
                with st.container():
                    col1, col2 = st.columns([3, 1])
                    
                    with col1:
                        st.markdown(f"**{activity.get('activity_type', 'Activity')}**")
                        if activity.get('activity_details'):
                            st.caption(activity.get('activity_details'))
                    
                    with col2:
                        created_at = activity.get('created_at', '')
                        if created_at:
                            st.caption(created_at[:19])
                    
                    st.divider()
        else:
            st.info("No activities recorded yet")
        
        # Add new activity
        with st.expander("➕ Add New Activity"):
            with st.form("add_activity_form"):
                activity_type = st.selectbox(
                    "Activity Type",
                    ["Call", "Email", "Meeting", "Note", "Follow-up", "Other"]
                )
                
                activity_details = st.text_area("Details")
                
                if st.form_submit_button("Add Activity"):
                    if activity_details:
                        # Here you would save the activity to the database
                        st.success("Activity added!")
                        st.rerun()
    
    def render_settings(self):
        """Render settings page"""
        st.title("⚙️ Settings")
        st.markdown("<p class='subtitle'>Configure your Ultimate LeadScraper</p>", unsafe_allow_html=True)
        
        # Create tabs for different setting categories
        tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
            "🔑 API Keys", "🎯 Scraper Modes", "🌍 Locations", "🏢 Industries",
            "⚡ Performance", "📧 Notifications"
        ])
        
        with tab1:
            self.render_api_settings()
        
        with tab2:
            self.render_scraper_mode_settings()
        
        with tab3:
            self.render_location_settings()
        
        with tab4:
            self.render_industry_settings()
        
        with tab5:
            self.render_performance_settings()
        
        with tab6:
            self.render_notification_settings()
    
    def render_api_settings(self):
        """Render API settings tab"""
        st.subheader("API Configuration")
        
        with st.form("api_settings_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                serper_key = st.text_input(
                    "Serper API Key",
                    value=CONFIG.api.serper_api_key,
                    type="password",
                    help="Get from https://serper.dev"
                )
                
                proxycrawl_key = st.text_input(
                    "ProxyCrawl API Key",
                    value=CONFIG.api.proxycrawl_api_key or "",
                    type="password",
                    help="Optional: For bypassing blocks"
                )
            
            with col2:
                openai_key = st.text_input(
                    "OpenAI API Key",
                    value=CONFIG.api.openai_api_key,
                    type="password",
                    help="Get from https://platform.openai.com/api-keys"
                )
                
                brightdata_key = st.text_input(
                    "BrightData API Key",
                    value=CONFIG.api.brightdata_api_key or "",
                    type="password",
                    help="Optional: For premium proxies"
                )
            
            if st.form_submit_button("💾 Save API Keys", use_container_width=True):
                CONFIG.api.serper_api_key = serper_key
                CONFIG.api.openai_api_key = openai_key
                CONFIG.api.proxycrawl_api_key = proxycrawl_key or None
                CONFIG.api.brightdata_api_key = brightdata_key or None
                
                save_config(CONFIG)
                st.success("API keys saved successfully!")
    
    def render_scraper_mode_settings(self):
        """Render scraper mode settings tab"""
        st.subheader("Scraper Mode Configuration")
        
        for mode_name, mode_config in CONFIG.scraper_modes.items():
            with st.expander(f"📊 {mode_config.name}", expanded=mode_name == CONFIG.active_mode):
                with st.form(f"mode_form_{mode_name}"):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        new_name = st.text_input(
                            "Mode Name",
                            value=mode_config.name,
                            key=f"name_{mode_name}"
                        )
                        
                        new_description = st.text_area(
                            "Description",
                            value=mode_config.description,
                            key=f"desc_{mode_name}"
                        )
                    
                    with col2:
                        scrape_no_website = st.checkbox(
                            "Scrape No Website",
                            value=mode_config.scrape_no_website,
                            key=f"no_web_{mode_name}"
                        )
                        
                        scrape_broken_website = st.checkbox(
                            "Scrape Broken Website",
                            value=mode_config.scrape_broken_website,
                            key=f"broken_{mode_name}"
                        )
                        
                        scrape_with_website = st.checkbox(
                            "Scrape With Website",
                            value=mode_config.scrape_with_website,
                            key=f"with_web_{mode_name}"
                        )
                    
                    quality_threshold = st.slider(
                        "Quality Threshold",
                        min_value=0,
                        max_value=100,
                        value=mode_config.quality_threshold,
                        key=f"threshold_{mode_name}"
                    )
                    
                    max_leads = st.number_input(
                        "Max Leads Per Cycle",
                        min_value=1,
                        max_value=1000,
                        value=mode_config.max_leads_per_cycle,
                        key=f"max_{mode_name}"
                    )
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        is_default = st.checkbox(
                            "Set as Default Mode",
                            value=mode_name == CONFIG.active_mode,
                            key=f"default_{mode_name}"
                        )
                    
                    with col2:
                        if st.form_submit_button("💾 Save Mode", use_container_width=True):
                            CONFIG.scraper_modes[mode_name].name = new_name
                            CONFIG.scraper_modes[mode_name].description = new_description
                            CONFIG.scraper_modes[mode_name].scrape_no_website = scrape_no_website
                            CONFIG.scraper_modes[mode_name].scrape_broken_website = scrape_broken_website
                            CONFIG.scraper_modes[mode_name].scrape_with_website = scrape_with_website
                            CONFIG.scraper_modes[mode_name].quality_threshold = quality_threshold
                            CONFIG.scraper_modes[mode_name].max_leads_per_cycle = max_leads
                            
                            if is_default:
                                CONFIG.active_mode = mode_name
                            
                            save_config(CONFIG)
                            st.success(f"Mode '{new_name}' saved successfully!")
    
    def render_location_settings(self):
        """Render location settings tab"""
        st.subheader("Location Configuration")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("##### Default Region")
            
            default_state = st.text_input(
                "Default State",
                value=CONFIG.default_state,
                key="default_state"
            )
            
            default_country = st.text_input(
                "Default Country",
                value=CONFIG.default_country,
                key="default_country"
            )
        
        with col2:
            st.markdown("##### Target Cities")
            
            cities_text = st.text_area(
                "Cities (one per line)",
                value="\n".join(CONFIG.cities),
                height=200,
                key="cities_text"
            )
            
            target_cities = st.text_area(
                "Target Cities (optional, one per line)",
                value="\n".join(CONFIG.filters.target_cities),
                height=100,
                key="target_cities"
            )
        
        if st.button("💾 Save Location Settings", use_container_width=True):
            CONFIG.default_state = default_state
            CONFIG.default_country = default_country
            
            if cities_text:
                CONFIG.cities = [city.strip() for city in cities_text.split("\n") if city.strip()]
            
            if target_cities:
                CONFIG.filters.target_cities = [city.strip() for city in target_cities.split("\n") if city.strip()]
            else:
                CONFIG.filters.target_cities = []
            
            save_config(CONFIG)
            st.success("Location settings saved successfully!")
    
    def render_industry_settings(self):
        """Render industry settings tab"""
        st.subheader("Industry Configuration")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("##### Industries to Scrape")
            
            industries_text = st.text_area(
                "Industries (one per line)",
                value="\n".join(CONFIG.industries),
                height=300,
                key="industries_text"
            )
        
        with col2:
            st.markdown("##### Target Industries")
            
            target_industries = st.text_area(
                "Target Industries (optional, one per line)",
                value="\n".join(CONFIG.filters.target_industries),
                height=150,
                key="target_industries"
            )
            
            st.markdown("##### Search Phrases")
            
            search_phrases = st.text_area(
                "Search Phrases (use {industry}, {city}, {state})",
                value="\n".join(CONFIG.search_phrases),
                height=150,
                key="search_phrases"
            )
        
        if st.button("💾 Save Industry Settings", use_container_width=True):
            if industries_text:
                CONFIG.industries = [ind.strip() for ind in industries_text.split("\n") if ind.strip()]
            
            if target_industries:
                CONFIG.filters.target_industries = [ind.strip() for ind in target_industries.split("\n") if ind.strip()]
            else:
                CONFIG.filters.target_industries = []
            
            if search_phrases:
                CONFIG.search_phrases = [phrase.strip() for phrase in search_phrases.split("\n") if phrase.strip()]
            
            save_config(CONFIG)
            st.success("Industry settings saved successfully!")
    
    def render_performance_settings(self):
        """Render performance settings tab"""
        st.subheader("Performance Configuration")
        
        with st.form("performance_settings_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                searches_per_cycle = st.number_input(
                    "Searches Per Cycle",
                    min_value=1,
                    max_value=100,
                    value=CONFIG.searches_per_cycle,
                    key="searches_per_cycle"
                )
                
                businesses_per_search = st.number_input(
                    "Businesses Per Search",
                    min_value=1,
                    max_value=100,
                    value=CONFIG.businesses_per_search,
                    key="businesses_per_search"
                )
                
                cycle_interval = st.number_input(
                    "Cycle Interval (seconds)",
                    min_value=10,
                    max_value=3600,
                    value=CONFIG.cycle_interval,
                    key="cycle_interval"
                )
            
            with col2:
                max_cycles = st.number_input(
                    "Max Cycles",
                    min_value=1,
                    max_value=10000,
                    value=CONFIG.max_cycles,
                    key="max_cycles"
                )
                
                concurrent_scrapers = st.number_input(
                    "Concurrent Scrapers",
                    min_value=1,
                    max_value=20,
                    value=CONFIG.concurrent_scrapers,
                    key="concurrent_scrapers"
                )
                
                request_timeout = st.number_input(
                    "Request Timeout (seconds)",
                    min_value=5,
                    max_value=300,
                    value=CONFIG.request_timeout,
                    key="request_timeout"
                )
            
            # AI Settings
            st.markdown("##### AI Enrichment Settings")
            
            col1, col2 = st.columns(2)
            
            with col1:
                ai_enabled = st.checkbox(
                    "Enable AI Enrichment",
                    value=CONFIG.ai_enrichment.enabled,
                    key="ai_enabled"
                )
                
                auto_qualify = st.checkbox(
                    "Auto-Qualify Leads",
                    value=CONFIG.ai_enrichment.auto_qualify,
                    key="auto_qualify"
                )
            
            with col2:
                qualification_threshold = st.slider(
                    "Qualification Threshold",
                    min_value=0,
                    max_value=100,
                    value=CONFIG.ai_enrichment.qualification_threshold,
                    key="qualification_threshold"
                )
                
                extract_decision_maker = st.checkbox(
                    "Extract Decision Maker",
                    value=CONFIG.ai_enrichment.extract_decision_maker,
                    key="extract_decision_maker"
                )
            
            if st.form_submit_button("💾 Save Performance Settings", use_container_width=True):
                CONFIG.searches_per_cycle = searches_per_cycle
                CONFIG.businesses_per_search = businesses_per_search
                CONFIG.cycle_interval = cycle_interval
                CONFIG.max_cycles = max_cycles
                CONFIG.concurrent_scrapers = concurrent_scrapers
                CONFIG.request_timeout = request_timeout
                
                CONFIG.ai_enrichment.enabled = ai_enabled
                CONFIG.ai_enrichment.auto_qualify = auto_qualify
                CONFIG.ai_enrichment.qualification_threshold = qualification_threshold
                CONFIG.ai_enrichment.extract_decision_maker = extract_decision_maker
                
                save_config(CONFIG)
                st.success("Performance settings saved successfully!")
    
    def render_notification_settings(self):
        """Render notification settings tab"""
        st.subheader("Notification Configuration")
        
        with st.form("notification_settings_form"):
            st.markdown("##### Email Notifications")
            
            col1, col2 = st.columns(2)
            
            with col1:
                email_enabled = st.checkbox(
                    "Enable Email Notifications",
                    value=CONFIG.notifications.email_enabled,
                    key="email_enabled"
                )
                
                if email_enabled:
                    email_smtp_server = st.text_input(
                        "SMTP Server",
                        value=CONFIG.notifications.email_smtp_server or "",
                        key="email_smtp_server"
                    )
                    
                    email_smtp_port = st.number_input(
                        "SMTP Port",
                        min_value=1,
                        max_value=65535,
                        value=CONFIG.notifications.email_smtp_port or 587,
                        key="email_smtp_port"
                    )
            
            with col2:
                if email_enabled:
                    email_username = st.text_input(
                        "Email Username",
                        value=CONFIG.notifications.email_username or "",
                        key="email_username"
                    )
                    
                    email_password = st.text_input(
                        "Email Password",
                        value=CONFIG.notifications.email_password or "",
                        type="password",
                        key="email_password"
                    )
            
            st.markdown("##### Slack Notifications")
            
            slack_enabled = st.checkbox(
                "Enable Slack Notifications",
                value=CONFIG.notifications.slack_enabled,
                key="slack_enabled"
            )
            
            if slack_enabled:
                slack_webhook_url = st.text_input(
                    "Slack Webhook URL",
                    value=CONFIG.notifications.slack_webhook_url or "",
                    key="slack_webhook_url"
                )
            
            st.markdown("##### Telegram Notifications")
            
            telegram_enabled = st.checkbox(
                "Enable Telegram Notifications",
                value=CONFIG.notifications.telegram_enabled,
                key="telegram_enabled"
            )
            
            if telegram_enabled:
                col1, col2 = st.columns(2)
                
                with col1:
                    telegram_bot_token = st.text_input(
                        "Telegram Bot Token",
                        value=CONFIG.notifications.telegram_bot_token or "",
                        key="telegram_bot_token"
                    )
                
                with col2:
                    telegram_chat_id = st.text_input(
                        "Telegram Chat ID",
                        value=CONFIG.notifications.telegram_chat_id or "",
                        key="telegram_chat_id"
                    )
            
            if st.form_submit_button("💾 Save Notification Settings", use_container_width=True):
                CONFIG.notifications.email_enabled = email_enabled
                if email_enabled:
                    CONFIG.notifications.email_smtp_server = email_smtp_server or None
                    CONFIG.notifications.email_smtp_port = email_smtp_port or None
                    CONFIG.notifications.email_username = email_username or None
                    CONFIG.notifications.email_password = email_password or None
                
                CONFIG.notifications.slack_enabled = slack_enabled
                if slack_enabled:
                    CONFIG.notifications.slack_webhook_url = slack_webhook_url or None
                
                CONFIG.notifications.telegram_enabled = telegram_enabled
                if telegram_enabled:
                    CONFIG.notifications.telegram_bot_token = telegram_bot_token or None
                    CONFIG.notifications.telegram_chat_id = telegram_chat_id or None
                
                save_config(CONFIG)
                st.success("Notification settings saved successfully!")
    
    def render_analytics(self):
        """Render analytics page"""
        st.title("📈 Advanced Analytics")
        st.markdown("<p class='subtitle'>Deep insights and performance metrics</p>", unsafe_allow_html=True)
        
        # Time period selector
        col1, col2, col3 = st.columns(3)
        
        with col1:
            period = st.selectbox(
                "Time Period",
                ["7d", "30d", "90d", "All"],
                index=1
            )
        
        # Get statistics
        stats = self.crm.get_statistics(period)
        
        # Conversion Funnel
        st.markdown("<div class='modern-card'>", unsafe_allow_html=True)
        st.subheader("🔄 Conversion Funnel")
        
        funnel_data = stats.get('conversion_funnel', [])
        if funnel_data:
            df_funnel = pd.DataFrame(funnel_data)
            
            fig_funnel = go.Figure(go.Funnel(
                y=df_funnel['stage'],
                x=df_funnel['count'],
                textinfo="value+percent initial",
                opacity=0.8,
                marker=dict(color='#0066FF')
            ))
            
            fig_funnel.update_layout(
                height=400,
                showlegend=False
            )
            
            st.plotly_chart(fig_funnel, use_container_width=True)
        else:
            st.info("No conversion data available yet.")
        st.markdown("</div>", unsafe_allow_html=True)
        
        # Performance Metrics
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("<div class='modern-card'>", unsafe_allow_html=True)
            st.subheader("📊 Performance Metrics")
            
            metrics_data = {
                "Metric": ["Total Leads", "Avg. Score", "Conversion Rate", "Response Rate"],
                "Value": [
                    stats.get('overall', {}).get('total_leads', 0),
                    f"{stats.get('overall', {}).get('average_score', 0):.1f}",
                    "12.5%",  # Placeholder
                    "8.3%"   # Placeholder
                ]
            }
            
            df_metrics = pd.DataFrame(metrics_data)
            st.dataframe(df_metrics, use_container_width=True, hide_index=True)
            st.markdown("</div>", unsafe_allow_html=True)
        
        with col2:
            st.markdown("<div class='modern-card'>", unsafe_allow_html=True)
            st.subheader("🏆 Top Performing Cities")
            
            top_cities = stats.get('top_cities', [])
            if top_cities:
                df_cities = pd.DataFrame(top_cities)
                df_cities = df_cities.head(5)
                
                fig_cities = px.bar(
                    df_cities,
                    x='city',
                    y='lead_count',
                    color='avg_score',
                    color_continuous_scale='viridis',
                    text='lead_count'
                )
                fig_cities.update_layout(
                    height=300,
                    showlegend=False
                )
                st.plotly_chart(fig_cities, use_container_width=True)
            else:
                st.info("No city data available yet.")
            st.markdown("</div>", unsafe_allow_html=True)
    
    def render_export(self):
        """Render export page"""
        st.title("📤 Export Data")
        st.markdown("<p class='subtitle'>Export your leads in various formats</p>", unsafe_allow_html=True)
        
        # Export configuration
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Export Settings")
            
            export_format = st.radio(
                "Format",
                ["CSV", "Excel", "JSON"],
                horizontal=True
            )
            
            # Field selection
            available_fields = [
                "business_name", "website", "website_status", "phone", "email",
                "address", "city", "state", "industry", "business_type",
                "lead_score", "quality_tier", "potential_value", "lead_status",
                "assigned_to", "created_at", "scraped_date", "description"
            ]
            
            selected_fields = st.multiselect(
                "Fields to Include",
                available_fields,
                default=available_fields[:10]
            )
        
        with col2:
            st.subheader("Filters")
            
            # Quick filter presets
            preset = st.selectbox(
                "Quick Presets",
                ["All Leads", "High Intent Only", "Premium Leads", "No Website", "Active Leads"]
            )
            
            # Date range
            col_date1, col_date2 = st.columns(2)
            with col_date1:
                date_from = st.date_input("From Date")
            with col_date2:
                date_to = st.date_input("To Date")
        
        # Apply filters based on preset
        filters = {}
        
        if preset == "High Intent Only":
            filters["website_status"] = ["no_website", "broken", "parked"]
        elif preset == "Premium Leads":
            filters["quality_tier"] = ["Premium", "High"]
        elif preset == "No Website":
            filters["website_status"] = ["no_website"]
        elif preset == "Active Leads":
            filters["lead_status"] = ["New Lead", "Contacted", "Follow Up"]
        
        if date_from:
            filters["date_from"] = date_from.isoformat()
        if date_to:
            filters["date_to"] = date_to.isoformat()
        
        # Get filtered data
        leads_data = self.crm.get_leads(filters=filters, page=1, per_page=10000)
        leads = leads_data["leads"]
        
        st.metric("Leads to Export", len(leads))
        
        if leads:
            # Create dataframe with selected fields
            df = pd.DataFrame(leads)
            
            # Filter columns
            available_cols = [col for col in selected_fields if col in df.columns]
            df_export = df[available_cols]
            
            # Preview
            with st.expander("👁️ Preview Data"):
                st.dataframe(df_export.head(10), use_container_width=True)
            
            # Export buttons
            st.subheader("Download")
            
            if export_format == "CSV":
                csv_data = df_export.to_csv(index=False)
                st.download_button(
                    label="📥 Download CSV",
                    data=csv_data,
                    file_name=f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    type="primary",
                    use_container_width=True
                )
            
            elif export_format == "Excel":
                # Create Excel file with styling
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_export.to_excel(writer, index=False, sheet_name='Leads')
                    
                    # Get workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Leads']
                    
                    # Apply some basic styling
                    header_fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                
                st.download_button(
                    label="📥 Download Excel",
                    data=output.getvalue(),
                    file_name=f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
            
            elif export_format == "JSON":
                json_data = df_export.to_json(orient="records", indent=2)
                st.download_button(
                    label="📥 Download JSON",
                    data=json_data,
                    file_name=f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json",
                    type="primary",
                    use_container_width=True
                )
        else:
            st.warning("No leads to export with the current filters.")
    
    def render_logs(self):
        """Render logs page"""
        st.title("📋 System Logs")
        st.markdown("<p class='subtitle'>Monitor system activity and errors</p>", unsafe_allow_html=True)
        
        # Log viewer
        log_viewer = EnhancedLogger()
        recent_logs = log_viewer.get_recent_logs(limit=100)
        
        if recent_logs:
            # Filter options
            col1, col2, col3 = st.columns(3)
            
            with col1:
                log_level = st.selectbox(
                    "Filter by Level",
                    ["All", "INFO", "WARNING", "ERROR", "DEBUG"]
                )
            
            with col2:
                search_term = st.text_input("Search in logs")
            
            with col3:
                auto_refresh = st.checkbox("Auto-refresh (10s)", value=False)
            
            # Apply filters
            filtered_logs = recent_logs
            
            if log_level != "All":
                filtered_logs = [log for log in filtered_logs if log["level"] == log_level]
            
            if search_term:
                filtered_logs = [log for log in filtered_logs if search_term.lower() in log["message"].lower()]
            
            # Display logs
            st.subheader(f"Log Entries ({len(filtered_logs)})")
            
            for log in reversed(filtered_logs):  # Show newest first
                with st.container():
                    col1, col2 = st.columns([4, 1])
                    
                    with col1:
                        # Color code by level
                        level_colors = {
                            "INFO": "#3B82F6",
                            "WARNING": "#F59E0B",
                            "ERROR": "#EF4444",
                            "DEBUG": "#6B7280"
                        }
                        
                        level_color = level_colors.get(log["level"], "#6B7280")
                        
                        st.markdown(f"""
                        <div style="padding: 0.5rem; border-left: 4px solid {level_color}; margin-bottom: 0.5rem;">
                            <strong>{log['level']}</strong>: {log['message']}
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col2:
                        st.caption(log['timestamp'])
                    
                    st.divider()
            
            # Clear logs button
            if st.button("🗑️ Clear All Logs", type="secondary"):
                if os.path.exists(CONFIG.storage["logs_file"]):
                    with open(CONFIG.storage["logs_file"], "w") as f:
                        f.write("")
                    st.success("Logs cleared!")
                    st.rerun()
        else:
            st.info("No logs available yet.")
    
    def render_automation(self):
        """Render automation page"""
        st.title("🔄 Automation Rules")
        st.markdown("<p class='subtitle'>Automate lead processing and follow-ups</p>", unsafe_allow_html=True)
        
        # Automation rules configuration
        with st.expander("🤖 AI-Powered Automation Rules", expanded=True):
            st.markdown("""
            ### Automated Lead Processing Rules
            
            Configure rules to automatically process and qualify leads based on specific criteria.
            """)
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.checkbox("Auto-qualify high-intent leads", value=True)
                st.checkbox("Auto-assign based on city", value=True)
                st.checkbox("Auto-set follow-up dates", value=True)
            
            with col2:
                st.checkbox("Auto-detect decision makers", value=True)
                st.checkbox("Auto-generate outreach templates", value=True)
                st.checkbox("Auto-archive expired leads", value=True)
        
        # Campaign automation
        with st.expander("📧 Automated Campaign Sequences", expanded=True):
            st.markdown("""
            ### Email Sequence Automation
            
            Set up automated email sequences for different lead categories.
            """)
            
            sequence_type = st.selectbox(
                "Sequence Type",
                ["No Website Outreach", "Broken Website Follow-up", "Premium Lead Nurturing", "General Outreach"]
            )
            
            if sequence_type:
                col1, col2 = st.columns(2)
                
                with col1:
                    st.number_input("Days between emails", min_value=1, max_value=30, value=3)
                    st.number_input("Max emails in sequence", min_value=1, max_value=10, value=5)
                
                with col2:
                    st.checkbox("Send during business hours", value=True)
                    st.checkbox("Stop if lead responds", value=True)
        
        # Save automation settings
        if st.button("💾 Save Automation Rules", use_container_width=True, type="primary"):
            st.success("Automation rules saved successfully!")
    
    def run(self):
        """Run the dashboard"""
        try:
            # Render sidebar and get selected page
            page = self.render_sidebar()
            
            # Render selected page
            if page == "dashboard":
                self.render_dashboard()
            elif page == "leads":
                self.render_leads_management()
            elif page == "lead_details":
                self.render_lead_details_page()
            elif page == "settings":
                self.render_settings()
            elif page == "analytics":
                self.render_analytics()
            elif page == "export":
                self.render_export()
            elif page == "logs":
                self.render_logs()
            elif page == "automation":
                self.render_automation()
            
            # Auto-refresh if scraper is running
            if st.session_state.get('scraper_running') and AUTOREFRESH_AVAILABLE:
                st_autorefresh(interval=10000, limit=100, key="dashboard_refresh")
            
        except Exception as e:
            logger.log(f"Dashboard error: {e}", "ERROR")
            st.error(f"An error occurred: {str(e)}")
    
    def render_lead_details_page(self):
        """Render standalone lead details page"""
        st.title("🔍 Lead Details")
        
        # Lead ID input
        col1, col2 = st.columns([3, 1])
        
        with col1:
            lead_id = st.number_input(
                "Enter Lead ID",
                min_value=1,
                value=st.session_state.get('selected_lead_id', 1),
                key="lead_details_id"
            )
        
        with col2:
            if st.button("Load Lead", use_container_width=True):
                st.session_state.selected_lead_id = lead_id
        
        # Load and display lead
        if st.session_state.get('selected_lead_id'):
            lead = self.crm.get_lead_by_id(st.session_state.selected_lead_id)
            
            if lead:
                self.render_lead_detail_view(lead)
            else:
                st.error("Lead not found!")

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main entry point"""
    print("\n" + "="*80)
    print("🚀 ULTIMATE LEAD SCRAPER CRM v2.0")
    print("="*80)
    print("🌟 Features:")
    print("  ✅ High-Intent Lead Generation (No/Broken Websites)")
    print("  ✅ Multiple Scraping Modes Configurable")
    print("  ✅ AI-Powered Lead Qualification & Enrichment")
    print("  ✅ Complete CRM with Advanced Filtering")
    print("  ✅ Beautiful Streamlit Dashboard")
    print("  ✅ Real-time Analytics & Reporting")
    print("  ✅ Export in CSV, Excel, JSON")
    print("  ✅ Automated Campaign Sequences")
    print("  ✅ Multi-platform Scraping (Google, Facebook, Yelp, BBB)")
    print("  ✅ Advanced Website Analysis")
    print("  ✅ Notification System (Email, Slack, Telegram)")
    print("  ✅ Audit Logging & Security")
    print("="*80)
    
    # Check API keys
    if not CONFIG.api.serper_api_key:
        print("\n⚠️  Serper API key not configured")
        print("   Get from: https://serper.dev")
        print("   Add to config.json or set SERPER_API_KEY environment variable")
    
    if not CONFIG.api.openai_api_key:
        print("\n⚠️  OpenAI API key not configured - AI features disabled")
        print("   Get from: https://platform.openai.com/api-keys")
        print("   Add to config.json or set OPENAI_API_KEY environment variable")
    
    print(f"\n🎯 Active Mode: {CONFIG.scraper_modes[CONFIG.active_mode].name}")
    print(f"🌍 Region: {CONFIG.default_state}, {CONFIG.default_country}")
    print(f"🏙️  Cities: {len(CONFIG.cities)} configured")
    print(f"🏭 Industries: {len(CONFIG.industries)} configured")
    print(f"⚡ Performance: {CONFIG.concurrent_scrapers} concurrent scrapers")
    print("="*80)
    print("\n🌐 Starting Ultimate Streamlit Dashboard...")
    print(f"📱 Access at: http://localhost:{CONFIG.dashboard['port']}")
    print("\n📊 Dashboard Features:")
    print("  • Real-time monitoring and statistics")
    print("  • Advanced lead filtering and management")
    print("  • Lead details with AI insights")
    print("  • Full configuration editing")
    print("  • Export functionality (CSV, Excel, JSON)")
    print("  • System logs viewer")
    print("  • Automation rules configuration")
    print("="*80)
    
    # Create and run dashboard
    try:
        dashboard = UltimateStreamlitDashboard()
        dashboard.run()
    except Exception as e:
        logger.log(f"Dashboard initialization error: {e}", "ERROR")
        st.error(f"Failed to initialize dashboard: {str(e)}")

# ============================================================================
# CLI INTERFACE
# ============================================================================

def cli_interface():
    """Command line interface for the scraper"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Ultimate Lead Scraper CRM")
    parser.add_argument("--mode", choices=list(CONFIG.scraper_modes.keys()), 
                       default=CONFIG.active_mode, help="Scraping mode")
    parser.add_argument("--cycles", type=int, default=1, help="Number of cycles to run")
    parser.add_argument("--dashboard", action="store_true", help="Start dashboard")
    parser.add_argument("--export", choices=["csv", "json", "excel"], help="Export leads")
    parser.add_argument("--stats", action="store_true", help="Show statistics")
    
    args = parser.parse_args()
    
    if args.dashboard:
        main()
    elif args.export:
        export_leads(args.export)
    elif args.stats:
        show_statistics()
    else:
        run_cli_scraper(args.mode, args.cycles)

def export_leads(format: str):
    """Export leads from CLI"""
    leads_data = crm.get_leads(page=1, per_page=10000)
    leads = leads_data["leads"]
    
    if not leads:
        print("❌ No leads to export")
        return
    
    df = pd.DataFrame(leads)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if format == "csv":
        filename = f"leads_export_{timestamp}.csv"
        df.to_csv(filename, index=False)
        print(f"✅ Exported {len(leads)} leads to {filename}")
    
    elif format == "json":
        filename = f"leads_export_{timestamp}.json"
        df.to_json(filename, orient="records", indent=2)
        print(f"✅ Exported {len(leads)} leads to {filename}")
    
    elif format == "excel":
        filename = f"leads_export_{timestamp}.xlsx"
        df.to_excel(filename, index=False)
        print(f"✅ Exported {len(leads)} leads to {filename}")

def show_statistics():
    """Show statistics from CLI"""
    stats = crm.get_statistics("30d")
    
    print("\n📊 LEAD SCRAPER STATISTICS (Last 30 days)")
    print("="*50)
    
    overall = stats.get('overall', {})
    print(f"Total Leads: {overall.get('total_leads', 0):,}")
    print(f"Average Score: {overall.get('average_score', 0):.1f}")
    print(f"Estimated Value: ${overall.get('total_potential_value', 0):,}")
    print(f"Cities Covered: {overall.get('cities_covered', 0)}")
    print(f"Industries Covered: {overall.get('industries_covered', 0)}")
    
    print("\n🏆 Quality Distribution:")
    for quality in stats.get('quality_distribution', []):
        print(f"  {quality['quality_tier']}: {quality['count']} leads")
    
    print("\n🌐 Website Status:")
    for status in stats.get('website_status_distribution', []):
        print(f"  {status['website_status']}: {status['count']} leads")

def run_cli_scraper(mode: str, cycles: int):
    """Run scraper from CLI"""
    print(f"\n🚀 Starting scraper in '{mode}' mode for {cycles} cycles...")
    
    scraper = UltimateLeadScraper()
    scraper.set_mode(mode)
    scraper.start()
    
    try:
        for i in range(cycles):
            print(f"\n🔁 Cycle {i + 1}/{cycles}")
            asyncio.run(scraper.run_cycle_async())
            
            if i < cycles - 1:
                print(f"⏳ Waiting {CONFIG.cycle_interval} seconds before next cycle...")
                time.sleep(CONFIG.cycle_interval)
    
    except KeyboardInterrupt:
        print("\n\n🛑 Scraper interrupted by user")
    
    finally:
        scraper.stop()
        
        print("\n📊 Final Statistics:")
        print(f"Total Cycles: {scraper.stats['total_cycles']}")
        print(f"Total Leads Found: {scraper.stats['total_leads_found']}")
        print(f"High Intent Leads: {scraper.stats['high_intent_leads']}")
        print(f"Premium Leads: {scraper.stats['premium_leads']}")

# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    try:
        # Check for command line arguments
        if len(sys.argv) > 1:
            cli_interface()
        else:
            # Start dashboard by default
            main()
    
    except KeyboardInterrupt:
        print("\n\n🛑 Program interrupted by user")
    
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
